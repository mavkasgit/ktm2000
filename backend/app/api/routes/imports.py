import json
from io import BytesIO
from decimal import Decimal
from pathlib import Path

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from openpyxl import Workbook
from pydantic import BaseModel, Field
from sqlalchemy import desc, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.models.import_template import ImportTemplate
from app.models.imports import ImportBatch, ImportBatchMode, ImportBatchStatus, ImportFile
from app.models.production_plan import PlanChangeSet, PlanPosition, ProductionPlan
from app.models.product import Product
from app.models.route import ProductionRoute, RouteRuleProfile
from app.models.techcard import Techcard
from app.services.import_normalization import default_import_normalization_rules, has_valid_normalization_rules
from app.services.plan_import_service import create_excel_import_change_set
from app.services.route_matcher import resolve_position_route

router = APIRouter(prefix="/imports", tags=["imports"])


class ImportPreviewOut(BaseModel):
    import_file_id: int
    import_batch_id: int
    production_plan_id: int
    change_set_id: int
    template_id: int | None = None
    rule_profile_id: int | None = None
    rules_snapshot: list[dict] = Field(default_factory=list)
    route_selection_diagnostics: dict = Field(default_factory=dict)
    sheet_name: str
    header_row_number: int
    summary: dict
    items: list[dict]


@router.post("/excel", response_model=ImportPreviewOut, status_code=status.HTTP_201_CREATED)
async def import_excel_plan(
    file: UploadFile = File(...),
    sheet_index: int = Form(0),
    mode: ImportBatchMode = Form(ImportBatchMode.create_plan),
    production_plan_id: int | None = Form(None),
    plan_month: str | None = Form(None),
    plan_version: str | None = Form(None),
    row_selection: str | None = Form(None),
    template_id: int | None = Query(None),
    column_mapping: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
) -> ImportPreviewOut:
    if template_id is None:
        raise HTTPException(status_code=400, detail="template_id is required")

    resolved_mapping = None
    resolved_normalization_rules = None
    rule_profile_id = None
    template = await db.get(ImportTemplate, template_id)
    if template is None:
        raise HTTPException(status_code=404, detail="Template not found")
    if not template.is_active:
        raise HTTPException(status_code=400, detail="Template is inactive")
    if not has_valid_normalization_rules(template.normalization_rules):
        raise HTTPException(status_code=400, detail="Template normalization rules are invalid")

    resolved_mapping = dict(template.column_mapping)
    resolved_normalization_rules = dict(template.normalization_rules)
    rule_profile_id = (
        await db.execute(
            select(RouteRuleProfile.id)
            .where(RouteRuleProfile.import_template_id == template_id)
            .order_by(desc(RouteRuleProfile.is_active), desc(RouteRuleProfile.priority), RouteRuleProfile.id.asc())
            .limit(1)
        )
    ).scalars().first()
    if column_mapping is not None:
        try:
            parsed_mapping = json.loads(column_mapping)
        except json.JSONDecodeError as exc:
            raise HTTPException(status_code=400, detail=f"Invalid column_mapping JSON: {exc}") from exc
        if not isinstance(parsed_mapping, dict):
            raise HTTPException(status_code=400, detail="column_mapping must be a JSON object")
        resolved_mapping = {**(resolved_mapping or {}), **parsed_mapping}

    content = await file.read()
    try:
        result = await create_excel_import_change_set(
            db,
            filename=file.filename or "workbook.xls",
            content=content,
            content_type=file.content_type,
            sheet_index=sheet_index,
            mode=mode,
            production_plan_id=production_plan_id,
            plan_month=plan_month,
            plan_version=plan_version,
            column_mapping=resolved_mapping,
            normalization_rules=resolved_normalization_rules,
            row_selection=row_selection,
            template_id=template_id,
            rule_profile_id=rule_profile_id,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    return ImportPreviewOut(**result)


class SheetListOut(BaseModel):
    sheets: list[str]


@router.post("/excel/sheets", response_model=SheetListOut)
async def list_excel_sheets(file: UploadFile = File(...)) -> SheetListOut:
    from io import BytesIO

    from app.services.excel_import import validate_excel_extension

    validate_excel_extension(file.filename or "")
    content = await file.read()
    from python_calamine import load_workbook

    workbook = load_workbook(BytesIO(content))
    return SheetListOut(sheets=workbook.sheet_names)


class SheetPreviewOut(BaseModel):
    sheet_name: str
    header_row_number: int
    total_rows: int
    summary: dict
    items: list[dict]


@router.post("/excel/preview", response_model=SheetPreviewOut)
async def preview_excel_sheet_endpoint(
    file: UploadFile = File(...),
    sheet_index: int = Form(0),
    row_selection: str | None = Form(None),
    template_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
) -> SheetPreviewOut:
    from app.services.plan_import_service import preview_excel_sheet

    resolved_mapping = None
    resolved_normalization_rules = None
    rule_profile_id = None
    if template_id is not None:
        template = await db.get(ImportTemplate, template_id)
        if template is None:
            raise HTTPException(status_code=404, detail="Template not found")
        if not template.is_active:
            raise HTTPException(status_code=400, detail="Template is inactive")
        if not has_valid_normalization_rules(template.normalization_rules):
            raise HTTPException(status_code=400, detail="Template normalization rules are invalid")
        resolved_mapping = dict(template.column_mapping)
        resolved_normalization_rules = dict(template.normalization_rules)
        rule_profile_id = (
            await db.execute(
                select(RouteRuleProfile.id)
                .where(RouteRuleProfile.import_template_id == template_id)
                .order_by(desc(RouteRuleProfile.is_active), desc(RouteRuleProfile.priority), RouteRuleProfile.id.asc())
                .limit(1)
            )
        ).scalars().first()

    content = await file.read()
    result = await preview_excel_sheet(
        db,
        filename=file.filename or "workbook.xls",
        content=content,
        sheet_index=sheet_index,
        column_mapping=resolved_mapping,
        normalization_rules=resolved_normalization_rules,
        row_selection=row_selection,
        rule_profile_id=rule_profile_id,
    )
    return SheetPreviewOut(**result)


def _test_workbook(
    *,
    sku: str = "ЮП-2630",
    product_name: str = "Стык с дюбелем 40мм 2,7 анод.серебро матовый",
    quantity: Decimal = Decimal("100"),
    comments: str | None = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "План май 26 05"
    ws.append(["", "", "Комментарий"])
    ws.append(["Заявка № 05", "май"])
    ws.append([])
    ws.append(["", "", "", "", "", "", "", "", "", "", "", "", "Формирование ящиков"])
    ws.append(
        [
            "Артикул",
            "пополнение",
            "Наименование",
            "остатки сырья на КТМ",
            "Цвет",
            "кол-во шт. в 2,7",
            "Длина, м",
            "Пробивка/сверловка",
            "Упаковка",
            "Примечание ",
            "Длина после упак, м",
            "кол-во штук готовой продукции",
            "Запад",
            "Восток",
            "Вид конечного продукта",
            "Комментарии",
        ]
    )
    qty = float(quantity)
    ws.append(
        [
            sku,
            "ТЗ",
            product_name,
            3400,
            "серебро",
            qty,
            2.7,
            "",
            "поф, красная этикетка РП 23*150 на каждый профиль и белая этикетка 58*30 на пачку из 10 шт",
            "",
            2.7,
            qty,
            qty,
            qty,
            "ГП",
            comments or "",
        ]
    )
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


@router.post("/excel/test", response_model=ImportPreviewOut, status_code=status.HTTP_201_CREATED)
async def import_test_excel(
    production_plan_id: int | None = Query(None),
    techcard_id: int | None = Query(None),
    run_id: str | None = Query(None),
    plan_month: str | None = Query(None),
    plan_version: str | None = Query(None),
    quantity: Decimal = Query(Decimal("100")),
    row_selection: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
) -> ImportPreviewOut:
    import logging
    logger = logging.getLogger(__name__)

    sku = "ЮП-2630"
    name = "Стык с дюбелем 40мм 2,7 анод.серебро матовый"
    if techcard_id is not None:
        techcard = await db.get(Techcard, techcard_id)
        if techcard is None:
            raise HTTPException(status_code=404, detail="Techcard not found")
        if techcard.product_id is None:
            raise HTTPException(status_code=400, detail="Techcard has no product_id")
        product = await db.get(Product, techcard.product_id)
        if product is None:
            raise HTTPException(status_code=404, detail="Techcard product not found")
        sku = product.sku
        name = product.name

    comments = f"TEST_RUN:{run_id}" if run_id else None
    content = _test_workbook(
        sku=sku,
        product_name=name,
        quantity=quantity,
        comments=comments,
    )
    try:
        result = await create_excel_import_change_set(
            db,
            filename=f"test-{sku}.xlsx",
            content=content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            sheet_index=0,
            mode=ImportBatchMode.append_to_plan if production_plan_id else ImportBatchMode.create_plan,
            production_plan_id=production_plan_id,
            plan_month=plan_month,
            plan_version=plan_version,
            column_mapping=None,
            normalization_rules=default_import_normalization_rules(),
            row_selection=row_selection,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except RuntimeError as exc:
        logger.exception("import_test_excel RuntimeError: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    except Exception as exc:
        logger.exception("import_test_excel unexpected error: %s", exc)
        raise HTTPException(status_code=500, detail=f"{type(exc).__name__}: {exc}") from exc
    return ImportPreviewOut(**result)


class ImportRecentOut(BaseModel):
    id: int
    production_plan_id: int
    change_set_id: int | None
    template_id: int | None = None
    rule_profile_id: int | None = None
    plan_name: str
    plan_no: str
    original_filename: str
    mode: str
    status: str
    sheet_name: str
    parsed_rows: int
    total_rows: int
    error_count: int
    warning_count: int
    summary: dict
    route_selection_diagnostics: dict
    created_at: str

    class Config:
        from_attributes = True


@router.get("/recent", response_model=list[ImportRecentOut])
async def list_recent_imports(
    limit: int = Query(10, ge=1, le=50),
    db: AsyncSession = Depends(get_db),
) -> list[ImportRecentOut]:
    result = await db.execute(
        select(ImportBatch, ProductionPlan, ImportFile, PlanChangeSet)
        .join(ProductionPlan, ImportBatch.production_plan_id == ProductionPlan.id)
        .join(ImportFile, ImportBatch.source_file_id == ImportFile.id)
        .outerjoin(PlanChangeSet, PlanChangeSet.import_batch_id == ImportBatch.id)
        .order_by(desc(ImportBatch.created_at))
        .limit(limit)
    )
    items = []
    for batch, plan, file, change_set in result.all():
        summary = batch.summary or {}
        items.append(
            ImportRecentOut(
                id=batch.id,
                production_plan_id=batch.production_plan_id,
                change_set_id=change_set.id if change_set else None,
                template_id=batch.template_id,
                rule_profile_id=batch.rule_profile_id,
                plan_name=plan.name,
                plan_no=plan.plan_no,
                original_filename=file.original_filename,
                mode=batch.mode.value,
                status=batch.status.value,
                sheet_name=batch.sheet_name,
                parsed_rows=batch.parsed_rows,
                total_rows=batch.total_rows,
                error_count=summary.get("error_count", 0),
                warning_count=summary.get("warning_count", 0),
                summary=summary,
                route_selection_diagnostics=batch.route_selection_diagnostics or {},
                created_at=batch.created_at.isoformat() if batch.created_at else "",
            )
        )
    return items


class ImportPositionOut(BaseModel):
    id: int
    source_row_number: int | None
    source_sku: str
    source_name: str | None
    quantity: str
    product_id: int | None
    product_name: str | None
    route_id: int | None
    route_name: str | None
    route_source: str | None
    route_origin: str | None
    route_match_quality: str | None
    route_match_reason: str | None
    route_assigned_at: str | None
    route_manual_confirmed_at: str | None
    status: str
    validation_status: str
    validation_errors: list
    import_batch_id: int | None


@router.get("/{batch_id}/positions", response_model=list[ImportPositionOut])
async def list_import_positions(batch_id: int, db: AsyncSession = Depends(get_db)) -> list[ImportPositionOut]:
    batch = await db.get(ImportBatch, batch_id)
    if batch is None:
        raise HTTPException(status_code=404, detail="Import batch not found")

    positions = (
        await db.execute(
            select(PlanPosition)
            .where(PlanPosition.import_batch_id == batch_id)
            .order_by(PlanPosition.source_row_number)
        )
    ).scalars().all()

    # Preload products for route lookup
    products_cache: dict[int, Product | None] = {}
    route_resolve_cache: dict[int, dict] = {}

    result = []
    for pos in positions:
        product = None
        if pos.product_id:
            if pos.product_id not in products_cache:
                products_cache[pos.product_id] = await db.get(Product, pos.product_id)
            product = products_cache[pos.product_id]

        if pos.id in route_resolve_cache:
            route_info = route_resolve_cache[pos.id]
        else:
            route_info = await resolve_position_route(db, pos)
            route_resolve_cache[pos.id] = route_info

        product_name = product.name if product else None

        result.append(
            ImportPositionOut(
                id=pos.id,
                source_row_number=pos.source_row_number,
                source_sku=pos.source_sku,
                source_name=pos.source_name,
                quantity=str(pos.quantity),
                product_id=pos.product_id,
                product_name=product_name,
                route_id=route_info.route_id,
                route_name=route_info.route_name,
                route_source=route_info.source,
                route_origin=route_info.route_origin,
                route_match_quality=route_info.route_match_quality,
                route_match_reason=route_info.route_match_reason,
                route_assigned_at=route_info.route_assigned_at.isoformat() if route_info.route_assigned_at else None,
                route_manual_confirmed_at=(
                    route_info.route_manual_confirmed_at.isoformat() if route_info.route_manual_confirmed_at else None
                ),
                status=pos.status.value,
                validation_status=pos.validation_status.value,
                validation_errors=pos.validation_errors or [],
                import_batch_id=pos.import_batch_id,
            )
        )

    return result


@router.get("/files/{file_id}/download")
async def download_import_file(file_id: int, db: AsyncSession = Depends(get_db)):
    from fastapi.responses import FileResponse
    from urllib.parse import quote

    file = await db.get(ImportFile, file_id)
    if file is None:
        raise HTTPException(status_code=404, detail="File not found")
    if not file.stored_path:
        raise HTTPException(status_code=404, detail="File content not available")

    path = Path(file.stored_path)
    if not path.exists():
        raise HTTPException(status_code=404, detail="File not found on disk")

    encoded_name = quote(file.original_filename)
    return FileResponse(
        path=path,
        filename=file.original_filename,
        media_type="application/octet-stream",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_name}",
        },
    )
