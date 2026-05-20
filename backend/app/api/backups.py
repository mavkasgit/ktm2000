import csv
import io
import json
import os
import shutil
import subprocess
import tempfile
import threading
import uuid
import zipfile
from datetime import datetime, timedelta
from pathlib import Path
from typing import Callable, Dict, List, Optional
from urllib.parse import unquote, urlparse

from fastapi import APIRouter, Body, Depends, Form, HTTPException, UploadFile, File, status
from fastapi.responses import FileResponse
from openpyxl import Workbook
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.database import get_db, engine

router = APIRouter(prefix="/backups", tags=["backups"])

BACKUPS_DIR = Path(settings.BACKUPS_PATH)
BACKUPS_DIR.mkdir(parents=True, exist_ok=True)

BACKUP_DUMP_NAME = "database.dump"
BACKUP_MANIFEST_NAME = "manifest.json"
BACKUP_TABLE_EXPORTS_DIR = "exports/tables"
BACKUP_TABLE_WORKBOOK_NAME = "ktm2000_tables.xlsx"
BACKUP_STORAGE_DIRS = {
    "imports": "IMPORT_STORAGE_DIR",
    "products": "PRODUCT_PHOTO_DIR",
}
BACKUP_JOBS: dict[str, dict] = {}
BACKUP_JOBS_LOCK = threading.Lock()


def _get_db_name() -> str:
    """Извлекает имя БД из DATABASE_URL."""
    url = settings.DATABASE_URL.replace("postgresql+asyncpg://", "postgresql://")
    return url.split("/")[-1]


def _get_db_connection(db_name: str | None = None) -> tuple[List[str], Dict[str, str]]:
    """Возвращает CLI-аргументы подключения и env для PostgreSQL tools."""
    parsed = urlparse(settings.DATABASE_URL.replace("postgresql+asyncpg://", "postgresql://"))
    user = unquote(parsed.username or "")
    password = unquote(parsed.password or "")
    host = parsed.hostname or "localhost"
    port = str(parsed.port or 5432)
    database = db_name or parsed.path.lstrip("/")

    env = os.environ.copy()
    env["PGCLIENTENCODING"] = "UTF8"
    if password:
        env["PGPASSWORD"] = password
    return ["-h", host, "-p", port, "-U", user, "-d", database], env


def _docker_container_available() -> bool:
    """Проверяет, доступен ли Docker и запущен ли контейнер PostgreSQL."""
    try:
        result = subprocess.run(
            ["docker", "ps", "--format", "{{.Names}}"],
            capture_output=True, text=True, check=False, timeout=10
        )
        return settings.POSTGRES_CONTAINER_NAME in result.stdout
    except Exception:
        return False


def _run_postgres_cmd_docker(cmd: List[str], db_name: str | None = None) -> subprocess.CompletedProcess:
    """Выполняет PostgreSQL-команду через docker exec внутрь контейнера БД."""
    if not _docker_container_available():
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail=f"PostgreSQL client tools недоступны, а контейнер {settings.POSTGRES_CONTAINER_NAME} не запущен.",
        )

    parsed = urlparse(settings.DATABASE_URL.replace("postgresql+asyncpg://", "postgresql://"))
    user = unquote(parsed.username or "")
    password = unquote(parsed.password or "")
    database = db_name or parsed.path.lstrip("/")

    docker_base = ["docker", "exec", "-i"]
    if password:
        docker_base.extend(["-e", f"PGPASSWORD={password}"])
    docker_base.append(settings.POSTGRES_CONTAINER_NAME)

    # Inside the container, PostgreSQL listens on localhost:5432
    # Strip any -h/-p/--host/--port flags from the original command
    args = list(cmd[1:])
    for flag in ["-h", "-p", "--host", "--port"]:
        while flag in args:
            idx = args.index(flag)
            args.pop(idx)
            if idx < len(args) and not args[idx].startswith("-"):
                args.pop(idx)

    tool = cmd[0]

    if tool == "pg_dump":
        filepath = None
        if "-f" in args:
            idx = args.index("-f")
            filepath = args[idx + 1]
            args.pop(idx)
            args.pop(idx)
        docker_cmd = docker_base + ["pg_dump"] + args + ["-U", user, "-d", database]
        result = subprocess.run(docker_cmd, capture_output=True, text=False, timeout=300)
        if result.returncode == 0 and filepath:
            Path(filepath).write_bytes(result.stdout)
        return subprocess.CompletedProcess(
            args=result.args,
            returncode=result.returncode,
            stdout=result.stdout.decode("utf-8", errors="replace") if result.stdout else "",
            stderr=result.stderr.decode("utf-8", errors="replace") if result.stderr else "",
        )

    if tool == "pg_restore":
        filepath = None
        if args and not args[-1].startswith("-"):
            filepath = args[-1]
            args = args[:-1]
        docker_cmd = docker_base + ["pg_restore"] + args + ["-U", user, "-d", database]
        if filepath:
            file_bytes = Path(filepath).read_bytes()
            result = subprocess.run(docker_cmd, input=file_bytes, capture_output=True, text=False, timeout=300)
        else:
            result = subprocess.run(docker_cmd, capture_output=True, text=False, timeout=300)
        return subprocess.CompletedProcess(
            args=result.args,
            returncode=result.returncode,
            stdout=result.stdout.decode("utf-8", errors="replace") if result.stdout else "",
            stderr=result.stderr.decode("utf-8", errors="replace") if result.stderr else "",
        )

    # psql and others
    docker_cmd = docker_base + [tool] + args + ["-U", user, "-d", database]
    result = subprocess.run(docker_cmd, capture_output=True, text=True, encoding="utf-8", errors="replace", timeout=300)
    return result


def _running_inside_docker() -> bool:
    """Проверяет, запущен ли процесс внутри Docker-контейнера."""
    return Path("/.dockerenv").exists()


def _run_postgres_cmd_local(cmd: List[str], db_name: str | None = None) -> subprocess.CompletedProcess:
    """Выполняет PostgreSQL-команду через локальные CLI-утилиты (psql/pg_dump/pg_restore) с подключением по TCP."""
    parsed = urlparse(settings.DATABASE_URL.replace("postgresql+asyncpg://", "postgresql://"))
    user = unquote(parsed.username or "")
    password = unquote(parsed.password or "")
    host = parsed.hostname or "localhost"
    port = str(parsed.port or 5432)
    database = db_name or parsed.path.lstrip("/")

    env = os.environ.copy()
    env["PGCLIENTENCODING"] = "UTF8"
    if password:
        env["PGPASSWORD"] = password

    full_cmd = cmd + ["-h", host, "-p", port, "-U", user, "-d", database]
    try:
        return subprocess.run(full_cmd, capture_output=True, text=True, encoding="utf-8", errors="replace", env=env, check=False, timeout=300)
    except subprocess.TimeoutExpired:
        raise HTTPException(
            status_code=status.HTTP_504_GATEWAY_TIMEOUT,
            detail=f"Команда {cmd[0]} превысила время ожидания (5 минут).",
        )


def _run_postgres_cmd(cmd: List[str], db_name: str | None = None) -> subprocess.CompletedProcess:
    """Выполняет pg_dump/pg_restore/psql.
    Внутри Docker-контейнера использует локальные CLI с TCP-подключением.
    На хост-машине сначала пробует локальные CLI, если недоступны — fallback через Docker exec.
    """
    if _running_inside_docker():
        return _run_postgres_cmd_local(cmd, db_name)

    connection_args, env = _get_db_connection(db_name)
    full_cmd = [cmd[0], *connection_args, *cmd[1:]]
    try:
        return subprocess.run(
            full_cmd,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            check=False,
            timeout=300,
            env=env,
        )
    except FileNotFoundError:
        return _run_postgres_cmd_docker(cmd, db_name)
    except subprocess.TimeoutExpired:
        raise HTTPException(
            status_code=status.HTTP_504_GATEWAY_TIMEOUT,
            detail=f"Команда {cmd[0]} превысила время ожидания (5 минут).",
        )


def _get_all_tables(db_name: str | None = None) -> List[str]:
    """Возвращает список всех пользовательских таблиц в public-схеме."""
    result = _run_postgres_cmd([
        "psql", "-t", "-A",
        "-c", "SELECT tablename FROM pg_tables WHERE schemaname='public' AND tablename NOT LIKE 'pg_%' AND tablename NOT LIKE 'sql_%' ORDER BY tablename;"
    ], db_name)
    if result.returncode != 0:
        return []
    tables = [line.strip() for line in result.stdout.strip().split("\n") if line.strip()]
    return tables


def _get_current_preview(db_name: str | None = None) -> Dict:
    """Анализ текущей БД: считает записи во всех таблицах."""
    tables = _get_all_tables(db_name)
    stats: Dict[str, int] = {}
    for table in tables:
        count_result = _run_postgres_cmd(["psql", "-t", "-A", "-c", f'SELECT COUNT(*) FROM "{table}";'], db_name)
        try:
            stats[table] = int(count_result.stdout.strip())
        except (ValueError, TypeError):
            stats[table] = 0
    return {
        "source_db": _get_db_name(),
        "backup_timestamp": datetime.now().isoformat(),
        "tables": stats,
        "storage": _storage_summary(),
    }


def _write_backup_meta(filename: str, meta: Dict) -> None:
    """Сохраняет JSON-метаданные рядом с бэкапом."""
    meta_path = BACKUPS_DIR / f"{filename}.json"
    meta_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")


def _read_backup_meta(filename: str) -> Optional[Dict]:
    """Читает JSON-метаданные бэкапа, если они есть."""
    meta_path = BACKUPS_DIR / f"{filename}.json"
    if meta_path.exists():
        try:
            return json.loads(meta_path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            return None
    return None


def _delete_backup_file(filename: str) -> None:
    """Удаляет файл бэкапа + JSON-метаданные."""
    for ext in ["", ".json"]:
        path = BACKUPS_DIR / f"{filename}{ext}"
        if path.exists():
            path.unlink()


def _iter_backup_files() -> list[Path]:
    return sorted(
        [*BACKUPS_DIR.glob("backup_*.zip"), *BACKUPS_DIR.glob("backup_*.dump")],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )


def _parse_backup_db_name(filename: str) -> str:
    stem = Path(filename).stem
    parts = stem.split("_")
    if len(parts) >= 4 and parts[0] == "backup":
        return "_".join(parts[1:-2])
    return "unknown"


def _is_archive_backup(path: Path) -> bool:
    return path.suffix.lower() == ".zip"


def _storage_roots() -> dict[str, Path]:
    return {name: Path(getattr(settings, key)) for name, key in BACKUP_STORAGE_DIRS.items()}


def _storage_summary() -> Dict[str, Dict[str, int]]:
    summary: Dict[str, Dict] = {}
    for name, root in _storage_roots().items():
        file_count = 0
        total_bytes = 0
        dir_count = 0
        folders: dict[str, dict[str, int | str]] = {}
        if root.exists():
            folders["."] = {"path": ".", "files": 0, "bytes": 0}
            for path in root.rglob("*"):
                if path.is_dir():
                    dir_count += 1
                    rel_dir = path.relative_to(root).as_posix()
                    folders.setdefault(rel_dir, {"path": rel_dir, "files": 0, "bytes": 0})
                    continue
                if path.is_file():
                    rel_parent = path.parent.relative_to(root).as_posix()
                    rel_parent = "." if rel_parent == "." else rel_parent
                    folder = folders.setdefault(rel_parent, {"path": rel_parent, "files": 0, "bytes": 0})
                    size = path.stat().st_size
                    file_count += 1
                    total_bytes += size
                    folder["files"] = int(folder["files"]) + 1
                    folder["bytes"] = int(folder["bytes"]) + size
        summary[name] = {
            "files": file_count,
            "bytes": total_bytes,
            "directories": dir_count,
            "folders": sorted(folders.values(), key=lambda folder: str(folder["path"])),
        }
    return summary


def _collect_storage_files() -> list[tuple[str, Path, Path]]:
    files: list[tuple[str, Path, Path]] = []
    for name, root in _storage_roots().items():
        if not root.exists():
            continue
        for path in root.rglob("*"):
            if path.is_file():
                files.append((name, root, path))
    return files


def _write_storage_dirs_to_zip(zip_file: zipfile.ZipFile, progress: Callable[[int, int, str], None] | None = None) -> None:
    files = _collect_storage_files()
    total = max(len(files), 1)
    for index, (name, root, path) in enumerate(files, start=1):
        zip_file.write(path, Path("data") / name / path.relative_to(root))
        if progress:
            progress(index, total, f"Архивация файлов: {name}/{path.relative_to(root).as_posix()}")


def _csv_export_filename(table_name: str) -> str:
    safe_name = "".join(char if char.isalnum() or char in "_-" else "_" for char in table_name)
    return f"{safe_name or 'table'}.csv"


def _worksheet_title(table_name: str, existing_titles: set[str]) -> str:
    invalid_chars = set('[]:*?/\\')
    title = "".join("_" if char in invalid_chars else char for char in table_name).strip() or "table"
    title = title[:31]
    candidate = title
    counter = 2
    while candidate in existing_titles:
        suffix = f"_{counter}"
        candidate = f"{title[:31 - len(suffix)]}{suffix}"
        counter += 1
    existing_titles.add(candidate)
    return candidate


def _quote_identifier(identifier: str) -> str:
    return '"' + identifier.replace('"', '""') + '"'


def _export_table_csv(table_name: str, db_name: str | None = None) -> str:
    result = _run_postgres_cmd(
        ["psql", "-c", f"COPY {_quote_identifier(table_name)} TO STDOUT WITH CSV HEADER;"],
        db_name,
    )
    if result.returncode != 0:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Ошибка CSV-экспорта таблицы {table_name}: {result.stderr}",
        )
    return result.stdout or ""


def _add_csv_to_workbook(workbook: Workbook, table_name: str, csv_data: str, existing_titles: set[str]) -> None:
    sheet = workbook.create_sheet(_worksheet_title(table_name, existing_titles))
    for row in csv.reader(io.StringIO(csv_data)):
        sheet.append(row)

    if sheet.max_column > 0:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        for column_cells in sheet.columns:
            max_len = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells[:100]:
                max_len = max(max_len, len(str(cell.value or "")))
            sheet.column_dimensions[column_letter].width = min(max(max_len + 2, 10), 48)


def _write_table_exports_to_zip(
    zip_file: zipfile.ZipFile,
    db_name: str,
    progress: Callable[[int, int, str], None] | None = None,
) -> list[dict[str, str]]:
    exports: list[dict[str, str]] = []
    tables = _get_all_tables(db_name)
    workbook = Workbook()
    workbook.remove(workbook.active)
    worksheet_titles: set[str] = set()
    total = max(len(tables), 1)
    for index, table_name in enumerate(tables, start=1):
        filename = _csv_export_filename(table_name)
        archive_path = f"{BACKUP_TABLE_EXPORTS_DIR}/{filename}"
        if progress:
            progress(index, total, f"Экспорт таблицы: {table_name}")
        csv_data = _export_table_csv(table_name, db_name)
        zip_file.writestr(archive_path, csv_data.encode("utf-8-sig"))
        _add_csv_to_workbook(workbook, table_name, csv_data, worksheet_titles)
        exports.append({"table": table_name, "path": archive_path, "format": "csv"})

    if not tables:
        sheet = workbook.create_sheet("README")
        sheet.append(["В базе данных не найдено пользовательских таблиц."])

    workbook_stream = io.BytesIO()
    workbook.save(workbook_stream)
    zip_file.writestr(BACKUP_TABLE_WORKBOOK_NAME, workbook_stream.getvalue())

    readme = (
        "Этот каталог содержит табличный экспорт для просмотра без восстановления БД.\n"
        "../ktm2000_tables.xlsx — все таблицы в одном Excel-файле, по листу на таблицу.\n"
        "tables/*.csv — отдельный CSV-файл на каждую таблицу в UTF-8 with BOM для Excel.\n"
        "Основной источник для восстановления системы: database.dump.\n"
        "Файлы приложения лежат в data/imports, data/products.\n"
    )
    zip_file.writestr("exports/README.txt", readme)
    return exports


def _read_zip_manifest(zip_path: Path) -> Optional[Dict]:
    try:
        with zipfile.ZipFile(zip_path, "r") as zip_file:
            with zip_file.open(BACKUP_MANIFEST_NAME) as manifest_file:
                return json.loads(manifest_file.read().decode("utf-8"))
    except (KeyError, OSError, zipfile.BadZipFile, json.JSONDecodeError, UnicodeDecodeError):
        return None


async def _save_upload_file(file: UploadFile, target_path: Path) -> None:
    with open(target_path, "wb") as target:
        while True:
            chunk = await file.read(1024 * 1024)
            if not chunk:
                break
            target.write(chunk)


def _extract_dump_from_archive(zip_path: Path, target_path: Path) -> None:
    try:
        with zipfile.ZipFile(zip_path, "r") as zip_file:
            with zip_file.open(BACKUP_DUMP_NAME) as dump_file:
                target_path.write_bytes(dump_file.read())
    except (KeyError, OSError, zipfile.BadZipFile) as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Архив бэкапа повреждён или не содержит {BACKUP_DUMP_NAME}: {exc}",
        )


def _extract_storage_dirs_from_archive(zip_path: Path, target_root: Path) -> None:
    allowed_prefixes = {f"data/{name}/": name for name in BACKUP_STORAGE_DIRS}
    try:
        with zipfile.ZipFile(zip_path, "r") as zip_file:
            for info in zip_file.infolist():
                if info.is_dir():
                    continue
                normalized = info.filename.replace("\\", "/")
                storage_name = None
                for prefix, name in allowed_prefixes.items():
                    if normalized.startswith(prefix):
                        storage_name = name
                        relative_name = normalized[len(prefix):]
                        break
                if storage_name is None or not relative_name:
                    continue

                relative_path = Path(relative_name)
                if relative_path.is_absolute() or ".." in relative_path.parts:
                    raise HTTPException(
                        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                        detail=f"Архив содержит небезопасный путь: {info.filename}",
                    )

                destination = target_root / storage_name / relative_path
                destination.parent.mkdir(parents=True, exist_ok=True)
                with zip_file.open(info) as source, open(destination, "wb") as target:
                    shutil.copyfileobj(source, target)
    except zipfile.BadZipFile as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Архив бэкапа повреждён: {exc}",
        )


def _replace_storage_dirs(extracted_root: Path) -> None:
    for name, target_root in _storage_roots().items():
        source_root = extracted_root / name
        target_root.mkdir(parents=True, exist_ok=True)

        for child in target_root.iterdir():
            if child.is_dir():
                shutil.rmtree(child)
            else:
                child.unlink()

        if not source_root.exists():
            continue
        for child in source_root.iterdir():
            target = target_root / child.name
            if child.is_dir():
                shutil.copytree(child, target)
            else:
                shutil.copy2(child, target)


def _preview_dump_file(dump_path: Path, source_db: str | None = None) -> Dict:
    preview_db = f"ktm2000_preview_{uuid.uuid4().hex[:8]}"
    try:
        create_result = _run_postgres_cmd(["psql", "-c", f'CREATE DATABASE "{preview_db}";'], "postgres")
        if create_result.returncode != 0:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Ошибка создания временной БД: {create_result.stderr}",
            )

        restore_result = _run_postgres_cmd(["pg_restore", str(dump_path)], preview_db)
        if restore_result.returncode not in (0, 1):
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"pg_restore ошибка: {restore_result.stderr}",
            )

        tables = _get_all_tables(preview_db)
        stats: Dict[str, int] = {}
        for table in tables:
            count_result = _run_postgres_cmd(["psql", "-t", "-A", "-c", f'SELECT COUNT(*) FROM "{table}";'], preview_db)
            try:
                stats[table] = int(count_result.stdout.strip())
            except (ValueError, TypeError):
                stats[table] = 0

        return {
            "source_db": source_db or _get_db_name(),
            "backup_timestamp": datetime.fromtimestamp(dump_path.stat().st_mtime).isoformat(),
            "tables": stats,
            "cached": False,
        }
    finally:
        _run_postgres_cmd(["psql", "-c", f'DROP DATABASE IF EXISTS "{preview_db}" WITH (FORCE);'], "postgres")


def _restore_dump_file(dump_path: Path, db_name: str) -> None:
    _assert_dump_readable(dump_path)

    safe_db_name = db_name.replace("'", "''")
    term_result = _run_postgres_cmd([
        "psql", "-c",
        f"SELECT pg_terminate_backend(pid) FROM pg_stat_activity WHERE datname = '{safe_db_name}' AND pid <> pg_backend_pid();"
    ], "postgres")

    reset_result = _run_postgres_cmd([
        "psql",
        "-v", "ON_ERROR_STOP=1",
        "-c", "DROP SCHEMA IF EXISTS public CASCADE; CREATE SCHEMA public;",
    ], db_name)
    if reset_result.returncode != 0:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Ошибка очистки БД перед restore: {reset_result.stderr}",
        )

    restore_result = _run_postgres_cmd([
        "pg_restore",
        "--no-owner",
        "--no-privileges",
        "--exit-on-error",
        str(dump_path),
    ], db_name)
    if restore_result.returncode != 0:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"pg_restore ошибка: {restore_result.stderr}",
        )


def _assert_dump_readable(dump_path: Path) -> None:
    result = _run_pg_restore_list(dump_path)
    if result.returncode != 0:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=(
                "Файл database.dump несовместим с текущей версией pg_restore. "
                "Создайте backup заново после обновления backend-контейнера. "
                f"pg_restore: {result.stderr}"
            ),
        )


def _run_pg_restore_list(dump_path: Path) -> subprocess.CompletedProcess:
    try:
        return subprocess.run(
            ["pg_restore", "--list", str(dump_path)],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            check=False,
            timeout=300,
        )
    except FileNotFoundError:
        if not _docker_container_available():
            raise HTTPException(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                detail="pg_restore недоступен, Docker fallback недоступен.",
            )
        file_bytes = dump_path.read_bytes()
        result = subprocess.run(
            ["docker", "exec", "-i", settings.POSTGRES_CONTAINER_NAME, "pg_restore", "--list"],
            input=file_bytes,
            capture_output=True,
            text=False,
            check=False,
            timeout=300,
        )
        return subprocess.CompletedProcess(
            args=result.args,
            returncode=result.returncode,
            stdout=result.stdout.decode("utf-8", errors="replace") if result.stdout else "",
            stderr=result.stderr.decode("utf-8", errors="replace") if result.stderr else "",
        )
    except subprocess.TimeoutExpired:
        raise HTTPException(
            status_code=status.HTTP_504_GATEWAY_TIMEOUT,
            detail="Проверка database.dump превысила время ожидания (5 минут).",
        )


def _run_alembic_upgrade() -> None:
    backend_dir = Path(__file__).resolve().parents[2]
    try:
        result = subprocess.run(
            ["alembic", "upgrade", "head"],
            cwd=str(backend_dir),
            capture_output=True,
            text=True,
            check=False,
            timeout=300,
        )
    except FileNotFoundError as exc:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail=f"Alembic недоступен для миграции после restore: {exc}",
        )
    except subprocess.TimeoutExpired:
        raise HTTPException(
            status_code=status.HTTP_504_GATEWAY_TIMEOUT,
            detail="Alembic migration превысила время ожидания (5 минут).",
        )

    if result.returncode != 0:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Alembic migration ошибка: {result.stderr}",
        )


def _set_job_progress(job_id: str, progress: int, stage: str, message: str, **extra) -> None:
    with BACKUP_JOBS_LOCK:
        job = BACKUP_JOBS.get(job_id)
        if not job:
            return
        job.update({
            "progress": max(0, min(100, progress)),
            "stage": stage,
            "message": message,
            "updated_at": datetime.now().isoformat(),
            **extra,
        })


def _create_backup_archive(job_id: str | None = None) -> Dict:
    def report(progress: int, stage: str, message: str, **extra) -> None:
        if job_id:
            _set_job_progress(job_id, progress, stage, message, **extra)

    report(5, "preparing", "Подготовка к созданию бэкапа")
    db_name = _get_db_name()
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"backup_{db_name}_{timestamp}.zip"
    filepath = BACKUPS_DIR / filename
    dump_path = BACKUPS_DIR / f"{filename}.{BACKUP_DUMP_NAME}"

    try:
        report(15, "dumping_database", "Создание database.dump")
        result = _run_postgres_cmd(["pg_dump", "-F", "c", "-f", str(dump_path)], db_name)
        if result.returncode != 0:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"pg_dump ошибка: {result.stderr}",
            )

        report(30, "analyzing", "Сбор статистики таблиц и файлов")
        meta = _get_current_preview(db_name)
        meta["filename"] = filename
        meta["format"] = "archive-v2"
        meta["comment"] = ""
        meta["storage"] = _storage_summary()

        with zipfile.ZipFile(filepath, "w", compression=zipfile.ZIP_DEFLATED) as zip_file:
            report(40, "writing_dump", "Добавление database.dump в архив")
            zip_file.write(dump_path, BACKUP_DUMP_NAME)

            def storage_progress(done: int, total: int, message: str) -> None:
                report(45 + int((done / total) * 25), "adding_files", message, files_done=done, files_total=total)

            _write_storage_dirs_to_zip(zip_file, storage_progress)

            def table_progress(done: int, total: int, message: str) -> None:
                report(72 + int((done / total) * 18), "exporting_tables", message, tables_done=done, tables_total=total)

            table_exports = _write_table_exports_to_zip(zip_file, db_name, table_progress)
            meta["table_exports"] = {
                "format": "csv+xlsx",
                "path": BACKUP_TABLE_EXPORTS_DIR,
                "workbook_path": BACKUP_TABLE_WORKBOOK_NAME,
                "tables": table_exports,
            }

            report(95, "writing_manifest", "Запись manifest.json")
            zip_file.writestr(BACKUP_MANIFEST_NAME, json.dumps(meta, ensure_ascii=False, indent=2))

        meta["size"] = filepath.stat().st_size
        _write_backup_meta(filename, meta)
    finally:
        if dump_path.exists():
            dump_path.unlink()

    result = {
        "filename": filename,
        "db_name": db_name,
        "size": filepath.stat().st_size,
        "created_at": datetime.now().isoformat(),
        "comment": "",
    }
    report(100, "completed", "Бэкап готов", result=result)
    return result


def _run_backup_job(job_id: str) -> None:
    try:
        result = _create_backup_archive(job_id)
        _set_job_progress(job_id, 100, "completed", "Бэкап готов", status="completed", result=result)
    except Exception as exc:
        detail = exc.detail if isinstance(exc, HTTPException) else str(exc)
        _set_job_progress(job_id, 100, "failed", "Ошибка создания бэкапа", status="failed", error=detail)


def _validate_admin(current_user: str = "admin") -> None:
    """Заглушка проверки прав администратора."""
    if current_user != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Недостаточно прав")


@router.get("/current-preview")
async def current_preview() -> Dict:
    """Мгновенный анализ текущей БД (без restore)."""
    _validate_admin()
    return _get_current_preview()


@router.post("", status_code=status.HTTP_201_CREATED)
async def create_backup() -> Dict:
    """Создать архивный бэкап текущей БД и файловых storage."""
    _validate_admin()
    return _create_backup_archive()


@router.post("/jobs", status_code=status.HTTP_202_ACCEPTED)
async def start_backup_job() -> Dict:
    """Запустить создание бэкапа в фоне и вернуть job_id для polling."""
    _validate_admin()
    job_id = uuid.uuid4().hex
    now = datetime.now().isoformat()
    with BACKUP_JOBS_LOCK:
        BACKUP_JOBS[job_id] = {
            "job_id": job_id,
            "status": "running",
            "stage": "queued",
            "message": "Задача поставлена в очередь",
            "progress": 0,
            "created_at": now,
            "updated_at": now,
            "result": None,
            "error": None,
        }
    thread = threading.Thread(target=_run_backup_job, args=(job_id,), daemon=True)
    thread.start()
    return BACKUP_JOBS[job_id]


@router.get("/jobs/{job_id}")
async def get_backup_job(job_id: str) -> Dict:
    _validate_admin()
    with BACKUP_JOBS_LOCK:
        job = BACKUP_JOBS.get(job_id)
        if not job:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Задача бэкапа не найдена")
        return dict(job)


@router.get("/config")
async def get_backup_config() -> Dict:
    """Получить текущее имя базы данных для подтверждения восстановления."""
    _validate_admin()
    return {"db_name": _get_db_name()}


@router.get("")
async def list_backups() -> List[Dict]:
    """Список всех бэкапов с метаданными из JSON."""
    _validate_admin()
    backups = []
    for f in _iter_backup_files():
        meta = _read_backup_meta(f.name)
        backups.append({
            "filename": f.name,
            "db_name": (meta or {}).get("source_db") or _parse_backup_db_name(f.name),
            "size": f.stat().st_size,
            "created_at": datetime.fromtimestamp(f.stat().st_mtime).isoformat(),
            "comment": meta.get("comment", "") if meta else "",
            "format": (meta or {}).get("format") or ("archive-v2" if _is_archive_backup(f) else "database-dump"),
        })
    return backups


@router.get("/{filename}/download")
async def download_backup(filename: str) -> FileResponse:
    """Скачать файл бэкапа."""
    _validate_admin()
    filepath = BACKUPS_DIR / filename
    if not filepath.exists() or not filepath.is_file():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Бэкап не найден")
    return FileResponse(
        path=str(filepath),
        filename=filename,
        media_type="application/octet-stream",
    )


@router.patch("/{filename}/comment")
async def update_backup_comment(filename: str, comment: str = Body(..., embed=True)) -> Dict:
    """Обновить комментарий к бэкапу."""
    _validate_admin()
    filepath = BACKUPS_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Бэкап не найден")

    meta = _read_backup_meta(filename) or {}
    meta["comment"] = comment
    _write_backup_meta(filename, meta)
    return {"filename": filename, "comment": comment}


@router.delete("/{filename}")
async def delete_backup(filename: str) -> Dict:
    """Удалить бэкап + JSON-метаданные."""
    _validate_admin()
    filepath = BACKUPS_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Бэкап не найден")

    _delete_backup_file(filename)
    return {"status": "deleted", "filename": filename}


@router.post("/bulk-delete")
async def bulk_delete(body: Dict) -> Dict:
    """Массовое удаление выбранных бэкапов."""
    _validate_admin()
    filenames = body.get("filenames", [])
    if not filenames:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Не выбрано ни одного файла")

    deleted = []
    not_found = []
    for filename in filenames:
        filepath = BACKUPS_DIR / filename
        if filepath.exists():
            _delete_backup_file(filename)
            deleted.append(filename)
        else:
            not_found.append(filename)

    return {"deleted": deleted, "not_found": not_found}


@router.post("/delete-older-than")
async def delete_older_than(body: Dict) -> Dict:
    """Удалить бэкапы старше указанного количества дней."""
    _validate_admin()
    days = body.get("days")
    if days is None or not isinstance(days, int) or days < 0:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Некорректное количество дней")

    cutoff = datetime.now() - timedelta(days=days)
    deleted = []
    not_found = []

    for f in _iter_backup_files():
        mtime = datetime.fromtimestamp(f.stat().st_mtime)
        if mtime < cutoff:
            _delete_backup_file(f.name)
            deleted.append(f.name)

    return {"deleted": deleted, "cutoff": cutoff.isoformat()}


@router.post("/{filename}/preview")
async def preview_backup(filename: str) -> Dict:
    """Получить статистику (превью) из существующего бэкапа.
    Сначала ищет JSON-кэш/manifest, если нет — делает restore во временную БД."""
    _validate_admin()
    filepath = BACKUPS_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Бэкап не найден")

    meta = _read_backup_meta(filename)
    if meta and "tables" in meta:
        return {
            "source_db": meta.get("source_db", _get_db_name()),
            "backup_timestamp": meta.get("backup_timestamp", datetime.fromtimestamp(filepath.stat().st_mtime).isoformat()),
            "tables": meta["tables"],
            "storage": meta.get("storage"),
            "table_exports": meta.get("table_exports"),
            "cached": True,
        }

    if _is_archive_backup(filepath):
        manifest = _read_zip_manifest(filepath)
        if manifest and "tables" in manifest:
            return {
                "source_db": manifest.get("source_db", _parse_backup_db_name(filename)),
                "backup_timestamp": manifest.get("backup_timestamp", datetime.fromtimestamp(filepath.stat().st_mtime).isoformat()),
                "tables": manifest["tables"],
                "storage": manifest.get("storage"),
                "table_exports": manifest.get("table_exports"),
                "cached": True,
            }
        with tempfile.TemporaryDirectory() as tmp_dir:
            dump_path = Path(tmp_dir) / BACKUP_DUMP_NAME
            _extract_dump_from_archive(filepath, dump_path)
            return _preview_dump_file(dump_path, _parse_backup_db_name(filename))

    return _preview_dump_file(filepath, _parse_backup_db_name(filename))


@router.post("/upload-preview")
async def upload_preview(file: UploadFile = File(...)) -> Dict:
    """Загрузить .dump/.zip файл и получить превью статистики."""
    _validate_admin()
    if not file.filename or not file.filename.lower().endswith((".dump", ".zip")):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Файл должен иметь расширение .dump или .zip"
        )

    suffix = Path(file.filename).suffix.lower()
    tmp_path = BACKUPS_DIR / f"upload_preview_{uuid.uuid4().hex[:8]}{suffix}"
    try:
        await _save_upload_file(file, tmp_path)

        uploaded_db_name = _parse_backup_db_name(file.filename)
        if _is_archive_backup(tmp_path):
            manifest = _read_zip_manifest(tmp_path)
            if manifest and "tables" in manifest:
                return {
                    "source_db": manifest.get("source_db", uploaded_db_name),
                    "backup_timestamp": manifest.get("backup_timestamp", datetime.fromtimestamp(tmp_path.stat().st_mtime).isoformat()),
                    "tables": manifest["tables"],
                    "storage": manifest.get("storage"),
                    "table_exports": manifest.get("table_exports"),
                    "cached": True,
                }
            with tempfile.TemporaryDirectory() as tmp_dir:
                dump_path = Path(tmp_dir) / BACKUP_DUMP_NAME
                _extract_dump_from_archive(tmp_path, dump_path)
                return _preview_dump_file(dump_path, uploaded_db_name)

        preview = _preview_dump_file(tmp_path, uploaded_db_name)
        if preview["source_db"] == "unknown":
            return {
                **preview,
                "source_db": uploaded_db_name,
            }
        return preview
    finally:
        if tmp_path.exists():
            tmp_path.unlink()


@router.post("/{filename}/restore")
async def restore_backup(filename: str, body: Dict) -> Dict:
    """Восстановить текущую БД и файлы из выбранного бэкапа."""
    _validate_admin()
    db_name = _get_db_name()

    if body.get("db_name") != db_name:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Для подтверждения введите точное имя базы данных: {db_name}"
        )

    filepath = BACKUPS_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Бэкап не найден")

    await engine.dispose()

    if _is_archive_backup(filepath):
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_root = Path(tmp_dir)
            dump_path = tmp_root / BACKUP_DUMP_NAME
            extracted_storage_root = tmp_root / "storage"
            _extract_dump_from_archive(filepath, dump_path)
            _extract_storage_dirs_from_archive(filepath, extracted_storage_root)
            _restore_dump_file(dump_path, db_name)
            _run_alembic_upgrade()
            _replace_storage_dirs(extracted_storage_root)
    else:
        _restore_dump_file(filepath, db_name)
        _run_alembic_upgrade()

    migrate_result = _run_postgres_cmd(["psql", "-c", "SELECT 1"], db_name)

    return {"status": "restored", "db_name": db_name, "filename": filename}


@router.post("/upload-restore")
async def upload_restore(file: UploadFile = File(...), confirmed_db_name: str = Form(...)) -> Dict:
    """Загрузить .dump/.zip файл и восстановить текущую БД и файлы из него."""
    _validate_admin()
    db_name = _get_db_name()

    if confirmed_db_name != db_name:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Для подтверждения введите точное имя базы данных: {db_name}"
        )

    if not file.filename or not file.filename.lower().endswith((".dump", ".zip")):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Файл должен иметь расширение .dump или .zip"
        )

    suffix = Path(file.filename).suffix.lower()
    tmp_filename = f"upload_restore_{uuid.uuid4().hex[:8]}{suffix}"
    tmp_path = BACKUPS_DIR / tmp_filename

    try:
        await _save_upload_file(file, tmp_path)

        await engine.dispose()

        if _is_archive_backup(tmp_path):
            with tempfile.TemporaryDirectory() as tmp_dir:
                tmp_root = Path(tmp_dir)
                dump_path = tmp_root / BACKUP_DUMP_NAME
                extracted_storage_root = tmp_root / "storage"
                _extract_dump_from_archive(tmp_path, dump_path)
                _extract_storage_dirs_from_archive(tmp_path, extracted_storage_root)
                _restore_dump_file(dump_path, db_name)
                _run_alembic_upgrade()
                _replace_storage_dirs(extracted_storage_root)
        else:
            _restore_dump_file(tmp_path, db_name)
            _run_alembic_upgrade()

        return {"status": "restored", "db_name": db_name, "filename": file.filename}
    finally:
        if tmp_path.exists():
            tmp_path.unlink()
