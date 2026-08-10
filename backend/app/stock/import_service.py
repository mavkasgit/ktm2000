"""Сервис для импорта остатков из Excel в Stock Ledger.

Содержит:
* ``parse_remainders_excel`` — парсинг Excel-файла с остатками.
* ``apply_remainders_import`` — создание ``StockTransaction`` через
  ``StockCommandService``, с опцией очистки существующих остатков.
* ``generate_remainders_template_for_location`` — генерация шаблона .xlsx.
"""
from __future__ import annotations

import json
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from functools import lru_cache
from io import BytesIO
from pathlib import Path
from typing import Literal

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.domain.dimensions import (
    DIMENSIONLESS_LABEL,
    LENGTH_MM,
    DimensionsValidationError,
    canonicalize_dimensions,
    format_dimensions,
    parse_length_m_to_mm,
)
from app.models.product import Product
from app.models.route import SectionOperation
from app.models.section import Section
from app.models.user import User
from app.stock.models import QualityState, Reason, StockBalance
from app.stock.services import StockCommand, StockCommandService
from app.services.dimension_validation import (
    MissingDimensionsError,
    resolve_product_dimensions,
)
from app.services.excel_import import parse_row_selection
from app.services.import_column_resolver import detect_header_row, resolve_columns
from app.services.route_storage_classifier import is_production_section

_OPERATIONS_COMMENT_RE = re.compile(r"операции:\s*([^|]+)", re.IGNORECASE)

# Дефолтный column_mapping остатков — источник заголовков, псевдонимов и
# позиций колонок (issue #15). Живёт в JSON-файле ``remainders_columns.json``,
# а не в таблице шаблонов импорта: это системный дефолт импорта остатков,
# а не пользовательский шаблон (ADR-0003, «Обновление»). Используется, когда
# вызывающий код не передал собственный mapping (старые прямые вызовы, тесты).


@lru_cache(maxsize=1)
def load_remainders_default_mapping() -> dict:
    """Загружает дефолтный ``column_mapping`` импорта остатков из JSON."""
    path = Path(__file__).with_name("remainders_columns.json")
    return json.loads(path.read_text(encoding="utf-8"))


_REMAINDERS_DEFAULT_MAPPING: dict = load_remainders_default_mapping()


def parse_operations_from_comment(comment: str | None) -> list[str]:
    """Извлекает названия операций из комментария транзакции импорта остатков."""
    if not comment:
        return []
    match = _OPERATIONS_COMMENT_RE.search(comment)
    if not match:
        return []
    return [part.strip() for part in re.split(r"[,;|]+", match.group(1).strip()) if part.strip()]


# ─── Data classes ──────────────────────────────────────────────────────────────


@dataclass
class RemainderItem:
    """Одна строка из Excel-файла с остатками после парсинга."""

    source_row_number: int
    sku: str
    quantity: float | None
    comment: str | None
    product_id: int | None
    product_name: str | None
    status: Literal["valid", "invalid"]
    errors: list[str]
    raw_values: list[str]
    # NEW: колонка «Выполненные операции» (сырая строка из Excel)
    completed_operations_raw: str | None = None
    # NEW: резолвнутые пройденные этапы (поднабор ops_dict)
    completed_stages: list[dict] | None = None
    # NEW: имя целевой секции из Excel
    target_section_name: str | None = None
    # NEW: ID целевой секции (None если не найдена или не указана)
    target_section_id: int | None = None
    # Сырое значение и резолвнутый статус качества из колонки «Статус качества»
    quality_state_raw: str | None = None
    quality_state: QualityState = QualityState.GOOD
    # NEW (ADR-0003, п. 3): сырое значение колонки «Длина» (метры, как в Excel)
    length_raw: str | None = None
    # NEW: габариты строки, например {"length_mm": 2700}; None = безразмерные
    dimensions: dict | None = None
    # NEW: подпись габарита для предпросмотра («2,7 м» / «—»)
    dimensions_label: str = DIMENSIONLESS_LABEL


@dataclass
class SheetSummary:
    """Сводка по результатам парсинга листа Excel."""

    total: int
    valid: int
    invalid: int
    quantity_total: float


@dataclass
class ImportResult:
    """Результат применения импорта остатков."""

    success: bool
    imported_count: int
    errors: list[str]
    transaction_ids: list[int]


# ─── Quality value aliases (не относятся к разрешению колонок) ────────────────

_FINAL_SCRAP_ALIASES = frozenset({
    "окончательный брак",
    "оконч брак",
    "окончательный",
    "оконч",
    "final scrap",
    "final_scrap",
})
_DEFECT_SCRAP_ALIASES = frozenset({
    "брак",
    "дефект",
    "defect",
    "scrap",
})
_QUALITY_VALUE_MAP: dict[str, QualityState] = {
    "good": QualityState.GOOD,
    "годный": QualityState.GOOD,
    "годные": QualityState.GOOD,
    "г": QualityState.GOOD,
    "ok": QualityState.GOOD,
}


def _norm_hdr(value: str) -> str:
    """Normalise a header string for comparison (lowercase, collapse whitespace)."""
    return " ".join(str(value).lower().replace("\xa0", " ").strip().split())


def _resolve_remainders_indices(
    column_mapping: dict,
    header_row: list[str] | None,
) -> tuple[
    int, int | None, int | None, int | None, int | None, int | None, int | None
]:
    """Field indices из column_mapping через единый резолвер (#15).

    Возвращает ``(sku_idx, qty_idx, comment_idx, completed_ops_idx,
    target_section_idx, quality_state_idx, length_idx)``. ``sku_idx`` по
    умолчанию 0, остальные — None, если колонка не найдена.
    """
    idx_map = resolve_columns(column_mapping, header_row)
    return (
        idx_map.get("sku", 0),
        idx_map.get("quantity"),
        idx_map.get("comment"),
        idx_map.get("completed_operations"),
        idx_map.get("target_section"),
        idx_map.get("quality_state"),
        idx_map.get("length"),
    )


def parse_quality_state_cell(
    value: str | None,
    *,
    default: QualityState = QualityState.GOOD,
) -> tuple[QualityState, str | None]:
    """Resolve a spreadsheet cell to ``QualityState``.

    Empty cells use ``default``. Unknown values return an error message.
    """
    if not value or value.strip() in ("", "—", "-"):
        return default, None
    norm = _norm_hdr(value)
    if norm in _FINAL_SCRAP_ALIASES:
        return QualityState.FINAL_SCRAP, None
    if norm in _DEFECT_SCRAP_ALIASES:
        return QualityState.SCRAP, None
    if norm in {"rework", "переделка", "доработка"}:
        return default, "Статус «Переделка» не поддерживается при импорте"
    resolved = _QUALITY_VALUE_MAP.get(norm)
    if resolved is not None:
        return resolved, None
    return default, f"Неизвестный статус качества: '{value.strip()}'"


def _parse_qty(value: object) -> Decimal | None:
    """Try to parse a cell value as a positive Decimal quantity."""
    if value is None or value == "" or value == 0:
        return None
    try:
        if isinstance(value, Decimal):
            qty = value
        elif isinstance(value, (int, float)):
            qty = Decimal(str(value))
        else:
            norm = str(value).replace(" ", "").replace(",", ".").strip()
            if not norm:
                return None
            qty = Decimal(norm)
        return qty if qty > 0 else None
    except (InvalidOperation, ValueError):
        return None


def _cell_txt(value: object) -> str:
    """Convert a cell value to its string representation."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _rows_to_cell_grid(raw_rows: list[list]) -> list[list[str]]:
    """Normalise raw cell rows to string grid."""
    return [[_cell_txt(c) for c in row] for row in raw_rows]


def _parse_remainders_grid(
    rows: list[list],
    row_selection: str | None = None,
    *,
    header_row_idx: int | None = None,
    column_mapping: dict | None = None,
) -> tuple[list[RemainderItem], SheetSummary]:
    """Parse a 2D grid of cells into remainder items.

    ``column_mapping`` — ``column_mapping`` шаблона (данные): заголовки,
    псевдонимы и позиции колонок. ``None`` → дефолт из JSON «маппинг остатков».
    """
    cell_rows = _rows_to_cell_grid(rows)
    if not cell_rows:
        raise ValueError("Нет данных для импорта")

    mapping = column_mapping if column_mapping is not None else _REMAINDERS_DEFAULT_MAPPING

    if header_row_idx is not None:
        resolved_header_idx: int | None = header_row_idx
        use_positional = False
    else:
        resolved_header_idx = detect_header_row(cell_rows, mapping)
        use_positional = resolved_header_idx is None

    if use_positional:
        (
            sku_idx,
            qty_idx,
            comment_idx,
            completed_ops_idx,
            target_section_idx,
            quality_state_idx,
            length_idx,
        ) = _resolve_remainders_indices(mapping, None)
    else:
        headers = cell_rows[resolved_header_idx]  # type: ignore[index]
        (
            sku_idx,
            qty_idx,
            comment_idx,
            completed_ops_idx,
            target_section_idx,
            quality_state_idx,
            length_idx,
        ) = _resolve_remainders_indices(mapping, headers)

    selected_rows: set[int] | None = None
    if row_selection:
        selected_rows = parse_row_selection(row_selection)

    data_rows: list[tuple[int, list[str]]] = []
    if selected_rows is not None:
        for rn in sorted(selected_rows):
            idx_0 = rn - 1
            if 0 <= idx_0 < len(cell_rows):
                if not use_positional and idx_0 == resolved_header_idx:
                    continue
                data_rows.append((rn, cell_rows[idx_0]))
    elif use_positional:
        for i, row in enumerate(cell_rows, start=1):
            data_rows.append((i, row))
    else:
        for i, row in enumerate(
            cell_rows[resolved_header_idx + 1 :],  # type: ignore[operator]
            start=resolved_header_idx + 2,  # type: ignore[operator]
        ):
            data_rows.append((i, row))

    items: list[RemainderItem] = []
    valid_count = 0
    invalid_count = 0
    quantity_total = 0.0

    for row_num, row in data_rows:
        raw = list(row)

        sku_raw = row[sku_idx] if sku_idx < len(row) else ""
        qty_val = (
            row[qty_idx]
            if qty_idx is not None and qty_idx < len(row)
            else None
        )
        comment_raw = (
            row[comment_idx]
            if comment_idx is not None and comment_idx < len(row)
            else ""
        )
        completed_ops_raw = (
            row[completed_ops_idx]
            if completed_ops_idx is not None and completed_ops_idx < len(row)
            else None
        ) or None
        target_section_name = (
            row[target_section_idx]
            if target_section_idx is not None and target_section_idx < len(row)
            else None
        ) or None
        quality_state_raw = (
            row[quality_state_idx]
            if quality_state_idx is not None and quality_state_idx < len(row)
            else None
        ) or None
        length_raw = (
            row[length_idx]
            if length_idx is not None and length_idx < len(row)
            else None
        ) or None

        sku = sku_raw.strip() if sku_raw else ""
        comment = comment_raw if comment_raw else None
        parsed_qty = _parse_qty(qty_val)
        if parsed_qty is None and sku:
            qty_missing_or_empty = (
                qty_idx is None
                or qty_idx >= len(row)
                or not _cell_txt(qty_val).strip()
            )
            if qty_idx is None or (use_positional and qty_missing_or_empty):
                parsed_qty = Decimal("1")

        has_raw_content = any(str(v).strip() for v in raw)

        errors: list[str] = []
        if not sku:
            if has_raw_content:
                errors.append("Не удалось распознать артикул в строке")
            else:
                errors.append("Не указан артикул")
        if parsed_qty is None:
            errors.append("Количество отсутствует, равно нулю или не является числом")

        resolved_quality, quality_error = parse_quality_state_cell(quality_state_raw)
        if quality_error:
            errors.append(quality_error)

        # Колонка «Длина»: метры («2,7») → мм (2700); пусто/«—» = не указана,
        # мусор — invalid строка (ADR-0003, п. 3).
        row_dimensions: dict | None = None
        if length_raw is not None and length_raw.strip() not in ("", "—", "-"):
            try:
                row_dimensions = {LENGTH_MM: parse_length_m_to_mm(length_raw)}
            except DimensionsValidationError as exc:
                errors.append(str(exc))

        if not sku and parsed_qty is None and not comment and not has_raw_content:
            continue

        item = RemainderItem(
            source_row_number=row_num,
            sku=sku,
            quantity=float(parsed_qty) if parsed_qty is not None else None,
            comment=comment,
            product_id=None,
            product_name=None,
            status="invalid" if errors else "valid",
            errors=errors,
            raw_values=raw,
            completed_operations_raw=completed_ops_raw,
            completed_stages=[],
            target_section_name=target_section_name,
            target_section_id=None,
            quality_state_raw=quality_state_raw,
            quality_state=resolved_quality,
            length_raw=length_raw,
            dimensions=row_dimensions,
            dimensions_label=format_dimensions(row_dimensions),
        )

        if errors:
            invalid_count += 1
        else:
            valid_count += 1
            if parsed_qty is not None:
                quantity_total += float(parsed_qty)

        items.append(item)

    summary = SheetSummary(
        total=len(items),
        valid=valid_count,
        invalid=invalid_count,
        quantity_total=round(quantity_total, 3),
    )
    return items, summary


_HEADER_SPLIT_MARKERS: tuple[tuple[str, str], ...] = (
    ("артикул", "выполненные операции"),
    ("артикул", "операции"),
    ("sku", "completed_operations"),
    ("sku", "операции"),
)


def _try_split_known_header(line: str) -> list[str] | None:
    """Split a concatenated header like «АртикулВыполненные операции»."""
    lower = line.lower().replace("\xa0", " ")
    for left, right in _HEADER_SPLIT_MARKERS:
        li = lower.find(left)
        ri = lower.find(right)
        if li >= 0 and ri > li:
            return [line[:ri].strip(), line[ri:].strip()]
    return None


def _try_split_sku_ops_line(line: str, sku_prefix: str | None = None) -> list[str] | None:
    """Split a concatenated data row «ЮП-460Окно, Дробеструй» into SKU + operations."""
    if sku_prefix and line.startswith(sku_prefix):
        rest = line[len(sku_prefix) :].strip()
        if rest:
            return [sku_prefix, rest]
    # SKU обычно заканчивается цифрой, операции начинаются с заглавной буквы
    match = re.match(r"^(.+?)(?<=\d)(?=[А-ЯЁA-Z])", line)
    if match:
        rest = line[match.end() :].strip()
        if rest:
            return [match.group(1), rest]
    match = re.match(r"^([\w./-]+?)(?=[А-ЯЁA-Z][а-яёa-z,])", line)
    if match:
        rest = line[match.end() :].strip()
        if rest:
            return [match.group(1), rest]
    return None


def _clipboard_text_to_rows(text: str) -> list[list[str]]:
    """Parse TSV/CSV clipboard text into a cell grid."""
    normalized = text.replace("\r\n", "\n").replace("\r", "\n").strip()
    if not normalized:
        raise ValueError("Буфер обмена пуст")

    lines = [line for line in normalized.split("\n") if line.strip()]
    if not lines:
        raise ValueError("Буфер обмена пуст")

    has_tabs = any("\t" in line for line in lines)
    has_semicolons = any(";" in line for line in lines) and not has_tabs

    if not has_tabs and not has_semicolons:
        split_rows: list[list[str]] = []
        sku_prefix: str | None = None
        for i, line in enumerate(lines):
            if i == 0:
                header_cells = _try_split_known_header(line)
                if header_cells:
                    split_rows.append(header_cells)
                    continue
            cells = _try_split_sku_ops_line(line, sku_prefix)
            if cells:
                if sku_prefix is None:
                    sku_prefix = cells[0]
                split_rows.append(cells)
                continue
            split_rows.append([line])
        if split_rows and any(len(r) > 1 for r in split_rows):
            return split_rows

    rows: list[list[str]] = []
    for line in lines:
        if "\t" in line:
            cells = line.split("\t")
        elif ";" in line:
            cells = line.split(";")
        elif re.search(r"  +", line):
            cells = re.split(r"  +", line)
        else:
            cells = [line]
        rows.append([cell.strip() for cell in cells])
    return rows


# ─── Core functions ────────────────────────────────────────────────────────────


async def parse_remainders_clipboard(
    text: str,
    row_selection: str | None = None,
    column_mapping: dict | None = None,
) -> tuple[str, int, list[RemainderItem], SheetSummary]:
    """Parse tab-separated clipboard data copied from Excel or a spreadsheet."""
    rows = _clipboard_text_to_rows(text)
    items, summary = _parse_remainders_grid(
        rows, row_selection, column_mapping=column_mapping
    )
    return "Буфер обмена", len(rows), items, summary


async def parse_remainders_excel(
    content: bytes,
    sheet_index: int = 0,
    row_selection: str | None = None,
    column_mapping: dict | None = None,
) -> tuple[str, int, list[RemainderItem], SheetSummary]:
    """Parse an Excel file with remainders data.

    Performs basic cell-level validation (SKU not empty, quantity > 0).
    Product lookup is **not** done here — call ``_lookup_products`` separately
    or use ``apply_remainders_import`` which does it internally.

    Args:
        content: Raw bytes of the .xlsx file.
        sheet_index: 0-based sheet index.
        row_selection: Optional selection string like ``"2-10,12"``.
        column_mapping: ``column_mapping`` шаблона (заголовки/псевдонимы/
            позиции колонок); ``None`` → дефолт из JSON «маппинг остатков».

    Returns:
        ``(sheet_name, total_rows, items, summary)``.

    Raises:
        ValueError: If the sheet index is out of range or the sheet is empty.
    """
    from python_calamine import load_workbook

    workbook = load_workbook(BytesIO(content))
    if sheet_index < 0 or sheet_index >= len(workbook.sheet_names):
        raise ValueError(
            f"Sheet index {sheet_index} not found, "
            f"available sheets: {len(workbook.sheet_names)}"
        )

    sheet = workbook.get_sheet_by_index(sheet_index)
    rows = list(sheet.iter_rows())
    if not rows:
        raise ValueError("Workbook sheet is empty")

    items, summary = _parse_remainders_grid(
        rows, row_selection, column_mapping=column_mapping
    )
    return sheet.name, len(rows), items, summary


async def _lookup_products(db: AsyncSession, items: list[RemainderItem]) -> None:
    """Batch-lookup products by SKU and update items in-place.

    Sets ``product_id``, ``product_name`` for found products.
    Marks valid items as ``invalid`` if their SKU is not found.
    """
    skus = [it.sku for it in items if it.status == "valid" and it.sku]
    if not skus:
        return

    result = await db.execute(select(Product).where(Product.sku.in_(skus)))
    products = {p.sku: p for p in result.scalars().all()}

    for item in items:
        prod = products.get(item.sku)
        if prod is not None:
            item.product_id = prod.id
            item.product_name = prod.name
        elif item.status == "valid":
            item.status = "invalid"
            item.errors.append(f"SKU '{item.sku}' not found in database")


def _missing_dimensions_message(missing_codes: list[str]) -> str:
    """Понятная ошибка предпросмотра для обязательных измерений без значения."""
    if missing_codes == [LENGTH_MM]:
        return (
            "Не указана длина: заполните колонку «Длина» "
            "или задайте типовой размер продукта в справочнике измерений"
        )
    codes = ", ".join(missing_codes)
    return (
        f"Не указаны обязательные измерения: {codes} — заполните их в файле "
        "или задайте типовой размер продукта в справочнике измерений"
    )


async def resolve_remainder_dimensions(
    db: AsyncSession,
    items: list[RemainderItem],
) -> None:
    """Резолв габаритов строк по справочнику измерений (ADR-0003, п. 3), in-place.

    Три ветки для каждой valid-строки с найденным продуктом:
    - колонка «Длина» заполнена → берём её (уже в ``item.dimensions``);
    - пустая → типовой размер продукта (``default_value`` привязки);
    - нет ни того, ни другого, а измерение ``is_required`` → строка invalid
      с понятной ошибкой в предпросмотре.

    Продукты без привязок (крепёж и пр.) остаются безразмерными
    (``dimensions=None``), если длина не указана явно.
    """
    for item in items:
        if item.status != "valid" or item.product_id is None:
            continue
        try:
            resolved = await resolve_product_dimensions(
                db, item.product_id, item.dimensions
            )
            item.dimensions = canonicalize_dimensions(resolved)
        except MissingDimensionsError as exc:
            item.status = "invalid"
            item.errors.append(_missing_dimensions_message(exc.missing_codes))
            continue
        except DimensionsValidationError as exc:
            item.status = "invalid"
            item.errors.append(str(exc))
            continue
        item.dimensions_label = format_dimensions(item.dimensions)


# ─── Operations & Section resolvers ──────────────────────────────────────────


async def resolve_operations_dictionary(db: AsyncSession) -> list[dict]:
    """Return the dictionary of significant production operations.

    Returns a list of dicts in ``RouteStepsDisplay`` format:
    ``[{sequence, section_code, section_name, section_icon, section_icon_color,
    operation_code, operation_name, op_icon, op_icon_color, is_significant}]``
    Only ``is_significant=True AND operation_type='production'`` are included.
    ``sequence`` is the route-level section order (``section.sort_order``), so
    operations from different sections never appear as «совмещено» in the UI.
    Ordered by ``section.sort_order, section.id, section_operation.sort_order, section_operation.id``.
    """
    stmt = (
        select(SectionOperation, Section)
        .join(Section, Section.id == SectionOperation.section_id)
        .where(
            SectionOperation.is_significant.is_(True),
            SectionOperation.operation_type == "production",
        )
        .order_by(Section.sort_order, Section.id, SectionOperation.sort_order, SectionOperation.id)
    )
    result = await db.execute(stmt)
    return [
        {
            "sequence": section.sort_order,
            "section_code": section.code,
            "section_name": section.name,
            "section_icon": section.icon,
            "section_icon_color": section.icon_color,
            "operation_code": so.operation_code,
            "operation_name": so.operation_name,
            "op_icon": so.icon,
            "op_icon_color": so.icon_color,
            "is_significant": so.is_significant,
        }
        for so, section in result.all()
    ]


async def resolve_completed_stages(
    db: AsyncSession,
    raw_ops_str: str | None,
    ops_dict: list[dict],
    errors: list[str] | None = None,
) -> list[dict]:
    """Parse ``raw_ops_str`` and match against ``ops_dict``.

    Splits by ``,`` / ``;`` / ``|``, trims, lowercases, and exact-matches against
    ``operation_name`` (normalised). Returns a subset of ``ops_dict`` preserving
    original order (``sequence`` from dictionary, not re-calculated).

    Edge cases:
    - Empty/``—`` input → empty list, no error.
    - Duplicate operation name in input → single match (dedup).
    - Operation not found in dictionary → optional warning appended to ``errors``,
      item is **not** invalidated.
    """
    if not raw_ops_str or raw_ops_str.strip() in ("", "—", "-"):
        return []

    parts = [p.strip().lower() for p in re.split(r"[,;|]+", raw_ops_str) if p.strip()]
    if not parts:
        return []

    matched: list[dict] = []
    seen_names: set[str] = set()

    for raw_name in parts:
        found = False
        for op in ops_dict:
            op_norm = op["operation_name"].strip().lower()
            if raw_name == op_norm and op_norm not in seen_names:
                matched.append(op)
                seen_names.add(op_norm)
                found = True
                break
        if not found and errors is not None:
            errors.append(f"Операция '{raw_name}' не найдена в справочнике")

    return matched


async def resolve_target_section(
    db: AsyncSession,
    name: str | None,
    item_errors: list[str] | None = None,
) -> tuple[int | None, str | None]:
    """Resolve a section name to ``(section_id, section_name)``.

    Match is case-insensitive, trimmed.  ``production``-type sections are rejected
    (remainders cannot be imported there).  Allowed types:
    ``raw_stock, wip_stock, finished_stock, scrap``.

    Returns ``(None, None)`` if not found or type is ``production``, with
    a warning appended to ``item_errors`` (if provided).
    """
    if not name or name.strip() in ("", "—", "-"):
        return (None, None)

    norm_name = name.strip()
    result = await db.execute(
        select(Section).where(Section.name.ilike(norm_name))
    )
    section = result.scalar_one_or_none()

    if section is None:
        if item_errors is not None:
            item_errors.append(f"Участок '{norm_name}' не найден")
        return (None, None)

    if is_production_section(section):
        if item_errors is not None:
            item_errors.append(
                f"Участок '{norm_name}' имеет тип production, "
                f"нельзя использовать как цель импорта"
            )
        return (None, None)

    return (section.id, section.name)


def _resolve_item_quality_state(
    item: RemainderItem,
    default_quality_state: QualityState,
    quality_state_overrides: dict[int, QualityState] | None,
) -> QualityState:
    if quality_state_overrides and item.source_row_number in quality_state_overrides:
        return quality_state_overrides[item.source_row_number]
    return item.quality_state or default_quality_state


async def apply_remainders_import(
    db: AsyncSession,
    location_id: int,
    items: list[RemainderItem],
    quality_state: QualityState = QualityState.GOOD,
    user: User | None = None,
    clear_existing: bool = False,
    skip_invalid: bool = True,
    target_section_overrides: dict[int, int] | None = None,
    quality_state_overrides: dict[int, QualityState] | None = None,
) -> ImportResult:
    """Apply remainders import: create ``StockTransaction`` records.

    This function:
    1. Validates the location exists.
    2. Looks up products by SKU (``_lookup_products``) and resolves per-row
       dimensions against the dimension dictionary
       (``resolve_remainder_dimensions``).
    3. If ``clear_existing=True``, zeros out current balances for
       ``(location_id, quality_state)`` with ``ADJUSTMENT_OUT``.
       **Cannot** be combined with ``target_section_overrides``.
    4. Creates ``MANUAL_IN`` transactions for each valid row.
       Uses per-row ``to_location_id`` from ``target_section_overrides``
       or ``item.target_section_id`` (fallback to ``location_id``).
       If ``item.completed_stages`` is non-empty, appends operation names
       to the transaction comment.
    5. Commits the database session.

    Args:
        db: Database session.
        location_id: Target ``Section`` id for import.
        items: Parsed ``RemainderItem`` list (product lookup done here).
        quality_state: Default quality state for rows without a column value.
        quality_state_overrides: Optional dict mapping ``source_row_number``
            to ``QualityState`` for per-row UI override.
        user: User performing the import; used for ``created_by`` fields.
        clear_existing: If ``True``, zero existing balances for the target
            ``(location, quality_state)`` before importing.
        skip_invalid: If ``True``, skip invalid rows and continue with valid
            ones; if ``False``, abort on any invalid row.
        target_section_overrides: Optional dict mapping ``source_row_number``
            to ``section_id`` for per-row target override.  Takes precedence
            over ``item.target_section_id``.

    Returns:
        ``ImportResult`` with success status, counts, errors, and transaction ids.
    """
    errors: list[str] = []
    transaction_ids: list[int] = []

    # --- Validate location ---------------------------------------------------
    location = await db.get(Section, location_id)
    if location is None:
        return ImportResult(
            success=False,
            imported_count=0,
            errors=[f"Location with id={location_id} not found"],
            transaction_ids=[],
        )

    # --- Look up products ----------------------------------------------------
    await _lookup_products(db, items)

    # --- Resolve dimensions (явная длина / типовой размер / invalid) ---------
    await resolve_remainder_dimensions(db, items)

    valid_items = [it for it in items if it.status == "valid" and it.product_id is not None]
    invalid_items = [it for it in items if it.status == "invalid"]

    for it in invalid_items:
        errors.append(f"Row {it.source_row_number}: {', '.join(it.errors)} (SKU={it.sku})")

    # --- Abort if not skipping invalid ---------------------------------------
    if not skip_invalid and invalid_items:
        return ImportResult(
            success=False,
            imported_count=0,
            errors=errors,
            transaction_ids=[],
        )

    # --- Check clear_existing + per-row override incompatibility ------------
    if clear_existing and target_section_overrides is not None:
        return ImportResult(
            success=False,
            imported_count=0,
            errors=["clear_existing не поддерживается при построчном указании участков"],
            transaction_ids=[],
        )

    svc = StockCommandService()

    # --- Clear existing balances if requested --------------------------------
    if clear_existing:
        qualities_to_clear = {
            _resolve_item_quality_state(item, quality_state, quality_state_overrides)
            for item in valid_items
        }
        rows = await db.execute(
            select(StockBalance).where(
                StockBalance.location_id == location_id,
                StockBalance.quality_state.in_(qualities_to_clear),
            )
        )
        for bal in rows.scalars().all():
            if bal.balance_qty > 0:
                cmd = StockCommand(
                    product_id=bal.product_id,
                    from_location_id=location_id,
                    quantity=bal.balance_qty,
                    # Гасим каждую габаритную группу отдельно (ADR-0001).
                    dimensions=bal.dimensions,
                    reason=Reason.ADJUSTMENT_OUT,
                    quality_state=bal.quality_state,
                    comment="Очистка перед импортом остатков",
                    source_ref="import_remainders_excel",
                    created_by=user.id if user else 1,
                    created_by_user_name=(
                        user.full_name
                        if (user and user.full_name)
                        else (user.username if user else "system")
                    ),
                )
                tx = await svc.record(db, cmd)
                transaction_ids.append(tx.id)

    # --- Import valid items --------------------------------------------------
    imported_count = 0
    for item in valid_items:
        # Determine per-row to_location_id
        to_loc: int = location_id
        if target_section_overrides:
            to_loc = target_section_overrides.get(
                item.source_row_number, to_loc
            )
        if to_loc == location_id:
            to_loc = item.target_section_id or location_id

        # Build comment with completed operations
        final_comment: str | None = item.comment
        if item.completed_stages:
            ops_names = ", ".join(
                s["operation_name"] for s in item.completed_stages
            )
            sep = " | " if final_comment else ""
            final_comment = f"{final_comment or ''}{sep}операции: {ops_names}"

        row_quality = _resolve_item_quality_state(
            item, quality_state, quality_state_overrides
        )
        cmd = StockCommand(
            product_id=item.product_id,  # type: ignore[arg-type]
            to_location_id=to_loc,
            quantity=Decimal(str(item.quantity)),
            # Габаритная группа строки (ADR-0003, п. 3); None = безразмерные.
            dimensions=item.dimensions,
            reason=Reason.MANUAL_IN,
            quality_state=row_quality,
            comment=final_comment or "Импорт остатков из Excel",
            source_ref="import_remainders_excel",
            created_by=user.id if user else 1,
            created_by_user_name=(
                user.full_name
                if (user and user.full_name)
                else (user.username if user else "system")
            ),
        )
        tx = await svc.record(db, cmd)
        transaction_ids.append(tx.id)
        imported_count += 1

    await db.commit()

    return ImportResult(
        success=True,
        imported_count=imported_count,
        errors=errors,
        transaction_ids=transaction_ids,
    )


_TEMPLATE_SECTION_FALLBACKS: dict[str, str] = {
    "RAW_STOCK": "Склад сырья",
    "PREP_STOCK": "Склад подготовки",
    "WIP_STOCK": "Склад полуфабриката",
}
_TEMPLATE_ROW3_OPERATION_NAMES = ("Дробеструй", "Чёрный", "Стрейч")


async def _template_section_name(db: AsyncSession, code: str) -> str:
    result = await db.execute(
        select(Section.name).where(Section.code == code, Section.is_active.is_(True))
    )
    name = result.scalar_one_or_none()
    return name or _TEMPLATE_SECTION_FALLBACKS[code]


async def _template_example_row3_operations(db: AsyncSession) -> str:
    ops_dict = await resolve_operations_dictionary(db)
    known = {op["operation_name"] for op in ops_dict}
    resolved = [name for name in _TEMPLATE_ROW3_OPERATION_NAMES if name in known]
    if resolved:
        return ", ".join(resolved)
    return ", ".join(_TEMPLATE_ROW3_OPERATION_NAMES)


async def generate_remainders_template_for_location(
    db: AsyncSession,
    location_id: int,
) -> bytes:
    """Generate an Excel template (.xlsx) for remainders import.

    Те же примеры строк, что в UI модалки импорта остатков:
    ``Артикул | Кол-во | Статус качества | Операции | Участок | Коммент. | Длина``
    («Длина» — в метрах, запятая или точка: «2,7»; пусто = типовой размер).

    Args:
        db: Database session.
        location_id: Location to validate existence.

    Returns:
        Raw bytes of the .xlsx file.

    Raises:
        ValueError: If the location does not exist.
    """
    from openpyxl import Workbook

    location = await db.get(Section, location_id)
    if location is None:
        raise ValueError(f"Location with id={location_id} not found")

    raw_name = await _template_section_name(db, "RAW_STOCK")
    prep_name = await _template_section_name(db, "PREP_STOCK")
    wip_name = await _template_section_name(db, "WIP_STOCK")
    row3_ops = await _template_example_row3_operations(db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Импорт остатков"

    ws.append(["Артикул", "Кол-во", "Статус качества", "Операции", "Участок", "Коммент.", "Длина"])
    ws.append(["361", 200, "Годный", "", raw_name, "", ""])
    ws.append(["ALS-1289", 150, "Годный", "Дробеструй", prep_name, "Партия A", "2,7"])
    ws.append(["ЮП-2630", 80, "Окончательный брак", row3_ops, wip_name, "Срочный заказ", "1,8"])

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 12

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─── Preview table query (server pagination / filter / sort) ───────────────────

REMAINDER_PREVIEW_SORT_FIELDS = frozenset({
    "row",
    "sku",
    "quantity",
    "length",
    "operations",
    "quality",
    "section",
    "errors",
})

_PREVIEW_QUALITY_LABELS: dict[QualityState, str] = {
    QualityState.GOOD: "Годный",
    QualityState.SCRAP: "Брак",
    QualityState.FINAL_SCRAP: "Окончательный брак",
    QualityState.REWORK: "Переделка",
}


@dataclass
class RemainderSectionMeta:
    """Compact per-row section info for client-side override validation."""

    source_row_number: int
    status: Literal["valid", "invalid"]
    target_section_id: int | None
    target_section_name: str | None


@dataclass
class RemainderPreviewQueryResult:
    items: list[RemainderItem]
    items_total: int
    section_meta: list[RemainderSectionMeta]


def _norm_preview_text(value: str) -> str:
    return " ".join(str(value).lower().split())


def _matches_preview_partial(haystack: str, needle: str | None) -> bool:
    if needle is None or not needle.strip():
        return True
    return _norm_preview_text(needle) in _norm_preview_text(haystack)


def _preview_has_section_in_file(item: RemainderItem) -> bool:
    name = (item.target_section_name or "").strip()
    return bool(name and name not in ("—", "-"))


def _preview_effective_section_id(
    item: RemainderItem,
    target_section_overrides: dict[int, int] | None,
) -> int | None:
    if target_section_overrides and item.source_row_number in target_section_overrides:
        return target_section_overrides[item.source_row_number]
    return item.target_section_id


def _preview_operations_label(item: RemainderItem) -> str:
    if item.completed_stages:
        return ", ".join(stage["operation_name"] for stage in item.completed_stages)
    raw = (item.completed_operations_raw or "").strip()
    return raw or "—"


def _preview_quality_label(
    item: RemainderItem,
    *,
    default_quality_state: QualityState,
    quality_state_overrides: dict[int, QualityState] | None,
) -> str:
    state = _resolve_item_quality_state(item, default_quality_state, quality_state_overrides)
    return _PREVIEW_QUALITY_LABELS.get(state, state.value)


def _preview_section_label(
    item: RemainderItem,
    *,
    target_section_overrides: dict[int, int] | None,
    section_names: dict[int, str],
) -> str:
    if _preview_has_section_in_file(item) and item.target_section_name:
        return item.target_section_name
    section_id = _preview_effective_section_id(item, target_section_overrides)
    if section_id is None:
        return "—"
    return section_names.get(section_id, f"#{section_id}")


def _preview_errors_label(item: RemainderItem) -> str:
    if not item.errors:
        return "—" if item.status == "valid" else "Ошибка"
    return ", ".join(item.errors)


def _preview_cell_value(
    item: RemainderItem,
    field: str,
    *,
    default_quality_state: QualityState,
    target_section_overrides: dict[int, int] | None,
    quality_state_overrides: dict[int, QualityState] | None,
    section_names: dict[int, str],
) -> str:
    if field == "row":
        return str(item.source_row_number)
    if field == "sku":
        return item.sku
    if field == "quantity":
        return "—" if item.quantity is None else str(item.quantity)
    if field == "length":
        return item.dimensions_label
    if field == "operations":
        return _preview_operations_label(item)
    if field == "quality":
        return _preview_quality_label(
            item,
            default_quality_state=default_quality_state,
            quality_state_overrides=quality_state_overrides,
        )
    if field == "section":
        return _preview_section_label(
            item,
            target_section_overrides=target_section_overrides,
            section_names=section_names,
        )
    if field == "errors":
        return _preview_errors_label(item)
    return ""


def _preview_matches_search(item: RemainderItem, search: str | None) -> bool:
    if search is None or not search.strip():
        return True
    needle = search.strip()
    haystacks = [
        item.sku,
        item.product_name or "",
        str(item.source_row_number),
    ]
    return any(_matches_preview_partial(haystack, needle) for haystack in haystacks)


def _preview_matches_column_filters(
    item: RemainderItem,
    *,
    default_quality_state: QualityState,
    target_section_overrides: dict[int, int] | None,
    quality_state_overrides: dict[int, QualityState] | None,
    section_names: dict[int, str],
    column_filters: dict[str, str | None],
) -> bool:
    for field, needle in column_filters.items():
        if needle is None or not needle.strip():
            continue
        cell = _preview_cell_value(
            item,
            field,
            default_quality_state=default_quality_state,
            target_section_overrides=target_section_overrides,
            quality_state_overrides=quality_state_overrides,
            section_names=section_names,
        )
        if not _matches_preview_partial(cell, needle):
            return False
    return True


def _preview_sort_key(
    item: RemainderItem,
    field: str,
    *,
    default_quality_state: QualityState,
    target_section_overrides: dict[int, int] | None,
    quality_state_overrides: dict[int, QualityState] | None,
    section_names: dict[int, str],
) -> tuple[int | float | str, ...]:
    if field == "row":
        return (item.source_row_number,)
    if field == "quantity":
        return (item.quantity if item.quantity is not None else -1,)
    if field == "length":
        # Сортируем по числовой длине; безразмерные (—) — в начале.
        length_mm = (item.dimensions or {}).get(LENGTH_MM)
        if isinstance(length_mm, (int, float)):
            return (float(length_mm),)
        return (-1.0,)
    cell = _preview_cell_value(
        item,
        field,
        default_quality_state=default_quality_state,
        target_section_overrides=target_section_overrides,
        quality_state_overrides=quality_state_overrides,
        section_names=section_names,
    )
    if field in {"row", "quantity"}:
        try:
            return (float(cell),)
        except ValueError:
            return (0.0,)
    return (cell.casefold(),)


def build_remainder_section_meta(items: list[RemainderItem]) -> list[RemainderSectionMeta]:
    return [
        RemainderSectionMeta(
            source_row_number=item.source_row_number,
            status=item.status,
            target_section_id=item.target_section_id,
            target_section_name=item.target_section_name,
        )
        for item in items
    ]


def query_remainder_preview_items(
    items: list[RemainderItem],
    *,
    search: str | None = None,
    filter_status: str = "all",
    sort_by: str = "row",
    sort_order: str = "asc",
    limit: int = 50,
    offset: int = 0,
    default_quality_state: QualityState = QualityState.GOOD,
    target_section_overrides: dict[int, int] | None = None,
    quality_state_overrides: dict[int, QualityState] | None = None,
    section_names: dict[int, str] | None = None,
    row: str | None = None,
    sku: str | None = None,
    quantity: str | None = None,
    length: str | None = None,
    operations: str | None = None,
    quality: str | None = None,
    section: str | None = None,
    errors: str | None = None,
) -> RemainderPreviewQueryResult:
    """Filter, sort and paginate parsed remainder preview rows in memory."""
    resolved_section_names = section_names or {}
    section_meta = build_remainder_section_meta(items)

    filtered = list(items)
    if filter_status == "invalid":
        filtered = [item for item in filtered if item.status == "invalid"]

    if search:
        filtered = [item for item in filtered if _preview_matches_search(item, search)]

    column_filters = {
        "row": row,
        "sku": sku,
        "quantity": quantity,
        "length": length,
        "operations": operations,
        "quality": quality,
        "section": section,
        "errors": errors,
    }
    if any(value and value.strip() for value in column_filters.values()):
        filtered = [
            item
            for item in filtered
            if _preview_matches_column_filters(
                item,
                default_quality_state=default_quality_state,
                target_section_overrides=target_section_overrides,
                quality_state_overrides=quality_state_overrides,
                section_names=resolved_section_names,
                column_filters=column_filters,
            )
        ]

    resolved_sort_by = sort_by if sort_by in REMAINDER_PREVIEW_SORT_FIELDS else "row"
    reverse = sort_order.lower() == "desc"
    filtered.sort(
        key=lambda item: _preview_sort_key(
            item,
            resolved_sort_by,
            default_quality_state=default_quality_state,
            target_section_overrides=target_section_overrides,
            quality_state_overrides=quality_state_overrides,
            section_names=resolved_section_names,
        ),
        reverse=reverse,
    )

    items_total = len(filtered)
    safe_limit = max(1, min(limit, 500))
    safe_offset = max(0, offset)
    page_items = filtered[safe_offset: safe_offset + safe_limit]

    return RemainderPreviewQueryResult(
        items=page_items,
        items_total=items_total,
        section_meta=section_meta,
    )
