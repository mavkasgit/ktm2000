"""Тесты issue #6 «Габариты в позиции плана» (ADR-0003).

Группа строк упаковочного плана = одна PlanPosition: строка с собственным
входом открывает группу, строка-продолжение (вход в merged-диапазоне или
тот же SKU с пустым входом) — ещё один выход той же операции.
"""
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import Workbook

from app.core.config import settings
from app.models.import_template import ImportTemplate
from app.models.production_plan import PlanPosition
from app.services.excel_import import parse_factory_plan_workbook

PLAN_HEADERS = [
    "Артикул",
    "пополнение",
    "Наименование",
    "остатки сырья на КТМ",
    "Цвет",
    "кол-во шт. в 2,7",
    "Длина, м",
    "Пробивка/сверловка",
    "Упаковка",
    "Примечание ",
    "Длина после упак, м",
    "кол-во штук готовой продукции",
    "Запад",
    "Восток",
    "Вид конечного продукта",
]

# 1-based индексы колонок входа в PLAN_HEADERS (для merge_cells).
COL_SKU = 1
COL_INPUT_QTY = 6
COL_INPUT_LEN = 7


def _plan_row(sku, name, input_qty, input_len, out_len, out_qty):
    return [sku, "ТЗ", name, 0, "", input_qty, input_len, "", "", "", out_len, out_qty, "", "", "ГП"]


def _plan_workbook(data_rows, merges=()):
    """Собрать xlsx в формате упаковочного плана; merges: (col, start_row, end_row)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "План май 26 05"
    ws.append(["", "", "Комментарий"])
    ws.append(["Заявка № 05", "май"])
    ws.append([])
    ws.append(["", "", "", "", "", "", "", "", "", "", "", "", "Формирование ящиков"])
    ws.append(PLAN_HEADERS)
    for row in data_rows:
        ws.append(row)
    for col, start_row, end_row in merges:
        ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def test_merged_input_group_is_single_position() -> None:
    """Кейс АТ-7121: merged-вход 150 × 2,7 на две строки → одна позиция
    с выходами 350 × 0,9 и 50 × 1,8."""
    content = _plan_workbook(
        [
            _plan_row("AT-7121", "Профиль АТ", 150, 2.7, 0.9, 350),
            _plan_row("", "", "", "", 1.8, 50),
        ],
        merges=[
            (COL_SKU, 6, 7),
            (COL_INPUT_QTY, 6, 7),
            (COL_INPUT_LEN, 6, 7),
        ],
    )
    parsed = parse_factory_plan_workbook(content, "plan.xlsx")

    assert len(parsed.parsed_rows) == 1
    group = parsed.parsed_rows[0]
    assert group.source_sku == "AT-7121"
    assert group.source_row_numbers == [6, 7]
    assert group.source_ref == "rows:6-7"
    assert group.quantity == 400
    assert group.input_quantity == 150
    assert group.input_dimensions == {"length_mm": 2700}
    assert [(o["quantity"], o["dimensions"]) for o in group.outputs] == [
        ("350", {"length_mm": 900}),
        ("50", {"length_mm": 1800}),
    ]
    # Баланс сходится: 150×2700 = 350×900 + 50×1800.
    assert not any(w.startswith("plan_group_balance_mismatch") for w in group.warnings)


def test_same_sku_empty_input_row_joins_group_without_merge() -> None:
    """Реальный файл: у АТ-7121 объединены только колонки наименования,
    вход продолжения просто пуст — группа собирается по эвристике
    «тот же SKU + пустой вход»."""
    content = _plan_workbook(
        [
            _plan_row("AT-7121", "Профиль АТ", 150, 2.7, 0.9, 350),
            _plan_row("AT-7121", "", "", "", 1.8, 50),
        ]
    )
    parsed = parse_factory_plan_workbook(content, "plan.xlsx")

    assert len(parsed.parsed_rows) == 1
    group = parsed.parsed_rows[0]
    assert group.source_row_numbers == [6, 7]
    assert group.quantity == 400
    assert len(group.outputs) == 2


def test_adjacent_same_sku_rows_with_own_input_stay_separate() -> None:
    content = _plan_workbook(
        [
            _plan_row("AT-7121", "Профиль АТ", 100, 2.7, 2.7, 100),
            _plan_row("AT-7121", "Профиль АТ", 200, 2.7, 2.7, 200),
        ]
    )
    parsed = parse_factory_plan_workbook(content, "plan.xlsx")

    assert len(parsed.parsed_rows) == 2
    assert parsed.parsed_rows[0].quantity == 100
    assert parsed.parsed_rows[1].quantity == 200
    assert all(len(row.outputs) == 1 for row in parsed.parsed_rows)


def test_rows_without_lengths_import_as_dimensionless() -> None:
    content = _plan_workbook([_plan_row("SKU-X", "Безразмерный", "", "", "", 500)])
    parsed = parse_factory_plan_workbook(content, "plan.xlsx")

    assert len(parsed.parsed_rows) == 1
    row = parsed.parsed_rows[0]
    assert row.quantity == 500
    assert row.input_quantity is None
    assert row.input_dimensions is None
    assert row.outputs == [{"row_number": 6, "quantity": "500", "dimensions": None}]
    assert not any(w.startswith("invalid_") for w in row.warnings)


def test_garbage_lengths_warn_but_do_not_fail() -> None:
    content = _plan_workbook(
        [
            _plan_row("SKU-Y", "Мусор во входе", 100, "abc", 0.9, 100),
            _plan_row("SKU-Z", "Мусор в выходе", 100, 2.7, "xyz", 100),
        ]
    )
    parsed = parse_factory_plan_workbook(content, "plan.xlsx")

    assert len(parsed.parsed_rows) == 2
    bad_input, bad_output = parsed.parsed_rows
    assert "invalid_input_length:row=6" in bad_input.warnings
    assert bad_input.input_dimensions is None
    assert "invalid_output_length:row=7" in bad_output.warnings
    assert bad_output.outputs[0]["dimensions"] is None


def test_empty_input_length_inferred_from_single_output() -> None:
    """Пустая «Длина, м» при заполненном выходе — вход без резки."""
    content = _plan_workbook([_plan_row("SKU-W", "Без резки", 100, "", 0.9, 100)])
    parsed = parse_factory_plan_workbook(content, "plan.xlsx")

    row = parsed.parsed_rows[0]
    assert row.input_dimensions == {"length_mm": 900}
    assert row.payload["input"]["inferred"] is True


def test_group_with_conflicting_outputs_resets_inferred_input() -> None:
    """Догадка «вход = выход» сбрасывается при втором выходе с другой длиной."""
    content = _plan_workbook(
        [
            _plan_row("SKU-V", "Резка без входной длины", 150, "", 0.9, 350),
            _plan_row("SKU-V", "", "", "", 1.8, 50),
        ]
    )
    parsed = parse_factory_plan_workbook(content, "plan.xlsx")

    assert len(parsed.parsed_rows) == 1
    group = parsed.parsed_rows[0]
    assert group.input_dimensions is None
    assert group.payload["input"]["inferred"] is False
    assert len(group.outputs) == 2


def test_unbalanced_group_gets_warning() -> None:
    content = _plan_workbook(
        [
            _plan_row("AT-7121", "Профиль АТ", 150, 2.7, 0.9, 100),
            _plan_row("AT-7121", "", "", "", 1.8, 50),
        ]
    )
    parsed = parse_factory_plan_workbook(content, "plan.xlsx")

    assert len(parsed.parsed_rows) == 1
    group = parsed.parsed_rows[0]
    assert any(w.startswith("plan_group_balance_mismatch") for w in group.warnings)


def test_group_fingerprint_is_idempotent_and_output_sensitive() -> None:
    rows = [
        _plan_row("AT-7121", "Профиль АТ", 150, 2.7, 0.9, 350),
        _plan_row("AT-7121", "", "", "", 1.8, 50),
    ]
    first = parse_factory_plan_workbook(_plan_workbook(rows), "plan.xlsx")
    second = parse_factory_plan_workbook(_plan_workbook(rows), "plan.xlsx")
    assert first.parsed_rows[0].source_fingerprint == second.parsed_rows[0].source_fingerprint

    changed_rows = [
        _plan_row("AT-7121", "Профиль АТ", 150, 2.7, 0.9, 350),
        _plan_row("AT-7121", "", "", "", 1.8, 60),
    ]
    changed = parse_factory_plan_workbook(_plan_workbook(changed_rows), "plan.xlsx")
    assert changed.parsed_rows[0].source_fingerprint != first.parsed_rows[0].source_fingerprint


async def _create_template(session, *, name: str, code: str) -> ImportTemplate:
    template = ImportTemplate(
        name=name,
        code=code,
        is_active=True,
        column_mapping={"sku": {"header": "Артикул", "column": "A"}},
    )
    session.add(template)
    await session.flush()
    return template


async def _seed_product_with_route(session, sku: str):
    from app.models.product import Product, ProductType
    from app.models.route import ProductionRoute, RouteOperation, RouteStage
    from app.models.section import Section
    from app.models.techcard import Techcard, TechcardLine

    product = Product(sku=sku, name=f"Product {sku}", type=ProductType.finished_good, unit="pcs")
    component = Product(sku=f"{sku}-RAW", name=f"Raw {sku}", type=ProductType.component, unit="pcs")
    sections = [Section(code="CUT", name="Cut"), Section(code="PACKING", name="Pack")]
    session.add_all([product, component, *sections])
    await session.flush()

    techcard = Techcard(product_id=product.id, version="v1", is_active=True)
    session.add(techcard)
    await session.flush()
    session.add(TechcardLine(techcard_id=techcard.id, component_product_id=component.id, quantity=1, unit="pcs"))

    route = ProductionRoute(name=f"Route {sku}", is_active=True)
    session.add(route)
    await session.flush()
    for index, section in enumerate(sections, start=1):
        stage = RouteStage(
            route_id=route.id,
            sequence=index * 10,
            section_id=section.id,
            is_final=index == len(sections),
        )
        session.add(stage)
        await session.flush()
        session.add(RouteOperation(route_stage_id=stage.id, sequence=1, operation_name=f"Step {index}"))
    return product


@pytest.mark.asyncio
async def test_group_import_apply_and_reimport_idempotency(client, session, tmp_path, monkeypatch) -> None:
    """Полный цикл: импорт группы → одна позиция с входом/выходами в БД и API,
    повторный импорт того же файла не дублирует позицию."""
    monkeypatch.setattr(settings, "IMPORT_STORAGE_DIR", str(tmp_path))
    await _seed_product_with_route(session, "AT-7121")
    template = await _create_template(session, name="Dims Template", code="dims-template")
    await session.commit()

    content = _plan_workbook(
        [
            _plan_row("AT-7121", "Профиль АТ", 150, 2.7, 0.9, 350),
            _plan_row("", "", "", "", 1.8, 50),
        ],
        merges=[
            (COL_SKU, 6, 7),
            (COL_INPUT_QTY, 6, 7),
            (COL_INPUT_LEN, 6, 7),
        ],
    )
    mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    response = await client.post(
        f"/api/imports/excel?template_id={template.id}",
        files={"file": ("plan.xlsx", content, mime)},
    )
    assert response.status_code == 201
    body = response.json()
    assert body["summary"]["total_positions"] == 1
    after = body["items"][0]["after_data"]
    assert after["input_quantity"] == "150"
    assert after["input_dimensions"] == {"length_mm": 2700}
    assert [(o["quantity"], o["dimensions"]) for o in after["outputs"]] == [
        ("350", {"length_mm": 900}),
        ("50", {"length_mm": 1800}),
    ]

    plan_id = body["production_plan_id"]
    apply_resp = await client.post(
        f"/api/production-plans/{plan_id}/change-sets/{body['change_set_id']}/apply"
    )
    assert apply_resp.status_code == 200
    assert apply_resp.json()["created_positions"] == 1

    from sqlalchemy import select

    position = (
        await session.execute(select(PlanPosition).where(PlanPosition.production_plan_id == plan_id))
    ).scalar_one()
    assert position.quantity == Decimal("400")
    assert position.input_quantity == Decimal("150")
    assert position.input_dimensions == {"length_mm": 2700}
    assert [(o["quantity"], o["dimensions"]) for o in position.outputs] == [
        ("350", {"length_mm": 900}),
        ("50", {"length_mm": 1800}),
    ]

    positions_resp = await client.get(f"/api/production-plans/{plan_id}/all-positions")
    assert positions_resp.status_code == 200
    api_position = positions_resp.json()[0]
    assert api_position["input_quantity"] == "150"
    assert api_position["input_dimensions"] == {"length_mm": 2700}
    assert len(api_position["outputs"]) == 2
    assert api_position["operation_summary"] == "150 шт × 2,7 м → 350 × 0,9 м + 50 × 1,8 м"

    # Повторный импорт того же файла — группа опознана по fingerprint.
    reimport = await client.post(
        f"/api/imports/excel?template_id={template.id}",
        data={"mode": "append_to_plan", "production_plan_id": str(plan_id)},
        files={"file": ("plan.xlsx", content, mime)},
    )
    assert reimport.status_code == 201
    reimport_body = reimport.json()
    assert [item["change_action"] for item in reimport_body["items"]] == ["ignore_unchanged"]

    apply2 = await client.post(
        f"/api/production-plans/{plan_id}/change-sets/{reimport_body['change_set_id']}/apply"
    )
    assert apply2.status_code == 200
    assert apply2.json()["created_positions"] == 0

    count = (
        await session.execute(select(PlanPosition).where(PlanPosition.production_plan_id == plan_id))
    ).scalars().all()
    assert len(count) == 1


@pytest.mark.asyncio
async def test_typical_size_fallback_from_product_dimensions(client, session, tmp_path, monkeypatch) -> None:
    """Группа без входной длины с разными выходами: вход берётся из типового
    размера продукта (справочник измерений)."""
    monkeypatch.setattr(settings, "IMPORT_STORAGE_DIR", str(tmp_path))
    from app.models.dimension import DimensionType, ProductDimension

    product = await _seed_product_with_route(session, "AT-7122")
    dimension_type = DimensionType(code="length_mm", name="Длина", unit="мм")
    session.add(dimension_type)
    await session.flush()
    session.add(
        ProductDimension(
            product_id=product.id,
            dimension_type_id=dimension_type.id,
            is_required=True,
            default_value=2700,
        )
    )
    template = await _create_template(session, name="Typical Template", code="typical-template")
    await session.commit()

    content = _plan_workbook(
        [
            _plan_row("AT-7122", "Профиль АТ2", 150, "", 0.9, 350),
            _plan_row("AT-7122", "", "", "", 1.8, 50),
        ]
    )
    response = await client.post(
        f"/api/imports/excel?template_id={template.id}",
        files={
            "file": ("plan.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        },
    )
    assert response.status_code == 201
    after = response.json()["items"][0]["after_data"]
    assert after["input_dimensions"] == {"length_mm": 2700}
