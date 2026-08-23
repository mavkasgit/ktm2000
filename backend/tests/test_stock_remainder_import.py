"""Тесты импорта остатков из Excel (Remainders Import).

Проверяют:
- Preview парсинга Excel (с валидацией, без записи в БД).
- Применение импорта (MANUAL_IN транзакции, обновление баланса).
- Пропуск невалидных строк (skip_invalid).
- Атомарный откат при skip_invalid=False.
- Очистку существующих остатков (clear_existing).
- Качество (quality_state) в создаваемых транзакциях.
- Обработку неизвестного SKU.
- Скачивание шаблона.
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO

import pytest
from httpx import AsyncClient
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Product, ProductType, Section
from app.models.import_template import ImportTemplate
from app.models.route import SectionOperation
from app.models.user import User, UserRole
from app.stock.import_service import parse_operations_from_comment
from app.stock.models import QualityState, Reason, StockBalance, StockTransaction
from app.stock.services import StockCommand, StockCommandService
from tests.test_integrity_invariants import assert_no_invariants_violations

pytestmark = pytest.mark.asyncio


# ─── Helpers ───────────────────────────────────────────────────────────────────


async def _make_product(
    session: AsyncSession,
    sku: str = "IMP-PROD",
    name: str | None = None,
) -> Product:
    product = Product(
        sku=sku,
        name=name or sku,
        type=ProductType.finished_good,
        unit="pcs",
        is_active=True,
    )
    session.add(product)
    await session.flush()
    return product


async def _make_location(session: AsyncSession, code: str = "STOCK-IMP") -> Section:
    section = Section(
        code=code,
        name=code,
        type="raw_stock",
        is_active=True,
        sort_order=0,
    )
    session.add(section)
    await session.flush()
    return section


def _make_excel(
    rows: list[tuple],
    headers: tuple[str, ...] = ("SKU", "Количество", "Комментарий"),
) -> BytesIO:
    """Create an .xlsx file in memory with the given header and data rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Остатки"
    ws.append(list(headers))
    for row in rows:
        ws.append(list(row))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _make_clipboard_tsv(
    rows: list[tuple],
    headers: tuple[str, ...] = ("SKU", "Количество", "Комментарий"),
) -> str:
    """Create TSV clipboard text copied from Excel."""
    lines = ["\t".join(headers)]
    for row in rows:
        lines.append("\t".join(str(cell) for cell in row))
    return "\n".join(lines)


async def _make_stock_balance(
    session: AsyncSession,
    product_id: int,
    location_id: int,
    qty: Decimal,
    quality_state: QualityState = QualityState.GOOD,
) -> None:
    """Create a stock balance directly via StockCommandService."""
    svc = StockCommandService()
    cmd = StockCommand(
        product_id=product_id,
        to_location_id=location_id,
        quantity=qty,
        reason=Reason.MANUAL_IN,
        quality_state=quality_state,
        created_by=1,
        created_by_user_name="system",
    )
    await svc.record(session, cmd)
    await session.commit()


async def _make_section_operation(
    session: AsyncSession,
    section: Section,
    operation_name: str = "Прессование",
    operation_code: str | None = None,
    is_significant: bool = True,
    operation_type: str = "production",
    sort_order: int = 0,
    icon: str | None = None,
    icon_color: str | None = None,
) -> SectionOperation:
    """Create a SectionOperation linked to the given section."""
    if operation_code is None:
        operation_code = operation_name.upper().replace(" ", "_")[:20]
    op = SectionOperation(
        section_id=section.id,
        operation_code=operation_code,
        operation_name=operation_name,
        is_significant=is_significant,
        operation_type=operation_type,
        sort_order=sort_order,
        icon=icon,
        icon_color=icon_color,
    )
    session.add(op)
    await session.flush()
    return op


async def _make_warehouse_section(
    session: AsyncSession,
    code: str = "WH-IMP",
    name: str | None = None,
    type: str = "wip_stock",
) -> Section:
    """Create a warehouse-type Section for import target testing."""
    section = Section(
        code=code,
        name=name or code,
        type=type,
        is_active=True,
        sort_order=0,
    )
    session.add(section)
    await session.flush()
    return section


# ─── Tests: Preview ────────────────────────────────────────────────────────────


async def test_preview_remainders_excel_happy_path(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """POST /import/remainders/preview возвращает items и summary."""
    product = await _make_product(session, "PREV-001", "Тестовый продукт 1")
    product2 = await _make_product(session, "PREV-002", "Тестовый продукт 2")
    location = await _make_location(session, "PREVIEW-LOC")
    await session.commit()

    excel_buf = _make_excel([
        ("PREV-001", 10, "коммент1"),
        ("PREV-002", 20, None),
    ])

    resp = await client.post(
        "/api/stock/import/remainders/preview",
        files={"file": ("test.xlsx", excel_buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={
            "location_id": str(location.id),
            "quality_state": "good",
            "sheet_index": "0",
        },
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()

    assert body["sheet_name"] == "Остатки"
    assert body["total_rows"] == 3  # header + 2 data rows
    assert body["summary"]["total"] == 2
    assert body["summary"]["valid"] == 2
    assert body["summary"]["invalid"] == 0
    assert body["summary"]["quantity_total"] == 30.0

    items = body["items"]
    assert len(items) == 2

    # Item 1
    assert items[0]["sku"] == "PREV-001"
    assert items[0]["quantity"] == 10.0
    assert items[0]["product_id"] == product.id
    assert items[0]["product_name"] == "Тестовый продукт 1"
    assert items[0]["status"] == "valid"
    assert items[0]["comment"] == "коммент1"
    assert items[0]["errors"] == []

    # Item 2
    assert items[1]["sku"] == "PREV-002"
    assert items[1]["quantity"] == 20.0
    assert items[1]["product_id"] == product2.id
    assert items[1]["status"] == "valid"
    assert items[1]["comment"] is None


async def test_preview_without_location_id(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """Preview работает без location_id — только парсинг и валидация строк."""
    await _make_product(session, "NO-LOC-001")
    await session.commit()

    excel_buf = _make_excel([("NO-LOC-001", 5, "")])
    resp = await client.post(
        "/api/stock/import/remainders/preview",
        files={"file": ("test.xlsx", excel_buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"sheet_index": "0"},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["summary"]["valid"] == 1
    assert body["items"][0]["sku"] == "NO-LOC-001"


async def test_preview_remainders_pagination_limit_offset(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """Preview возвращает items_total и страницу items."""
    for idx in range(1, 6):
        await _make_product(session, f"PAGE-{idx:03d}")
    await session.commit()

    rows = [(f"PAGE-{idx:03d}", idx, None) for idx in range(1, 6)]
    excel_buf = _make_excel(rows)

    resp = await client.post(
        "/api/stock/import/remainders/preview",
        files={"file": ("test.xlsx", excel_buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"sheet_index": "0", "limit": "2", "offset": "0", "sort_by": "row", "sort_order": "asc"},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["summary"]["total"] == 5
    assert body["items_total"] == 5
    assert body["limit"] == 2
    assert body["offset"] == 0
    assert len(body["items"]) == 2
    assert len(body["section_meta"]) == 5
    assert body["items"][0]["sku"] == "PAGE-001"
    assert body["items"][1]["sku"] == "PAGE-002"

    resp_page2 = await client.post(
        "/api/stock/import/remainders/preview",
        files={"file": ("test.xlsx", excel_buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"sheet_index": "0", "limit": "2", "offset": "2", "sort_by": "row", "sort_order": "asc"},
    )
    assert resp_page2.status_code == 200, resp_page2.text
    body2 = resp_page2.json()
    assert body2["items_total"] == 5
    assert len(body2["items"]) == 2
    assert body2["items"][0]["sku"] == "PAGE-003"


async def test_preview_remainders_search_and_invalid_filter(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """Preview search и filter_status=invalid работают на сервере."""
    await _make_product(session, "FILTER-OK")
    await session.commit()

    excel_buf = _make_excel([
        ("FILTER-OK", 10, "ok"),
        ("FILTER-MISSING", 5, "bad"),
    ])

    search_resp = await client.post(
        "/api/stock/import/remainders/preview",
        files={"file": ("test.xlsx", excel_buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"sheet_index": "0", "search": "FILTER-OK", "limit": "50"},
    )
    assert search_resp.status_code == 200, search_resp.text
    search_body = search_resp.json()
    assert search_body["items_total"] == 1
    assert search_body["items"][0]["sku"] == "FILTER-OK"
    assert search_body["summary"]["total"] == 2

    invalid_resp = await client.post(
        "/api/stock/import/remainders/preview",
        files={"file": ("test.xlsx", excel_buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"sheet_index": "0", "filter_status": "invalid", "limit": "50"},
    )
    assert invalid_resp.status_code == 200, invalid_resp.text
    invalid_body = invalid_resp.json()
    assert invalid_body["items_total"] == 1
    assert invalid_body["items"][0]["status"] == "invalid"
    assert invalid_body["summary"]["invalid"] == 1


async def test_preview_remainders_excel_unknown_sku(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """Неизвестный SKU → invalid строка в preview."""
    location = await _make_location(session, "PREVIEW-UNK")
    await session.commit()

    excel_buf = _make_excel([
        ("UNKNOWN-SKU", 5, None),
    ])

    resp = await client.post(
        "/api/stock/import/remainders/preview",
        files={"file": ("test.xlsx", excel_buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"location_id": str(location.id), "sheet_index": "0"},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()

    assert body["summary"]["total"] == 1
    assert body["summary"]["valid"] == 0
    assert body["summary"]["invalid"] == 1
    items = body["items"]
    assert items[0]["status"] == "invalid"
    assert any("not found" in e for e in items[0]["errors"])


# ─── Tests: Import ─────────────────────────────────────────────────────────────


async def test_import_remainders_excel_creates_manual_in_transactions(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """Happy path: создаются StockTransaction с reason=MANUAL_IN и баланс обновлён."""
    product = await _make_product(session, "IMPORT-001")
    location = await _make_location(session, "IMPORT-LOC")
    await session.commit()

    excel_buf = _make_excel([
        ("IMPORT-001", 100, "тестовый импорт"),
    ])

    resp = await client.post(
        "/api/stock/import/remainders",
        files={"file": ("test.xlsx", excel_buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={
            "location_id": str(location.id),
            "quality_state": "good",
            "sheet_index": "0",
            "skip_invalid": "true",
            "clear_existing": "false",
        },
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["success"] is True
    assert body["imported_count"] == 1
    assert len(body["transaction_ids"]) == 1

    # Проверяем транзакцию в БД
    tx_id = body["transaction_ids"][0]
    tx = await session.get(StockTransaction, tx_id)
    assert tx is not None
    assert tx.product_id == product.id
    assert tx.to_location_id == location.id
    assert tx.from_location_id is None
    assert tx.reason == Reason.MANUAL_IN
    assert float(tx.quantity) == 100.0
    assert tx.source_ref.startswith("import_remainders:")

    # Проверяем баланс
    balance = await session.execute(
        select(StockBalance).where(
            StockBalance.product_id == product.id,
            StockBalance.location_id == location.id,
        )
    )
    bal = balance.scalar_one_or_none()
    assert bal is not None
    assert float(bal.balance_qty) == 100.0


async def test_import_remainders_excel_skip_invalid(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """1 валидная + 1 невалидная (пустой SKU) → imported_count=1, errors есть."""
    product = await _make_product(session, "SKIP-VALID")
    location = await _make_location(session, "SKIP-LOC")
    await session.commit()

    excel_buf = _make_excel([
        ("SKIP-VALID", 50, "ок"),
        ("", 10, "нет SKU"),
    ])

    resp = await client.post(
        "/api/stock/import/remainders",
        files={"file": ("test.xlsx", excel_buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={
            "location_id": str(location.id),
            "skip_invalid": "true",
        },
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["success"] is True
    assert body["imported_count"] == 1
    assert len(body["errors"]) == 1  # одна ошибка о пустом SKU
    assert "артикул" in body["errors"][0].lower()

    # Только 1 транзакция создана
    txs = (await session.execute(select(StockTransaction))).scalars().all()
    assert len(txs) == 1


async def test_import_remainders_excel_atomic_fail(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """skip_invalid=False с невалидной строкой → success=false, ничего не записано."""
    product = await _make_product(session, "ATOMIC-OK")
    location = await _make_location(session, "ATOMIC-LOC")
    await session.commit()

    excel_buf = _make_excel([
        ("ATOMIC-OK", 30, "хорошая строка"),
        ("", None, "плохая строка"),
    ])

    resp = await client.post(
        "/api/stock/import/remainders",
        files={"file": ("test.xlsx", excel_buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={
            "location_id": str(location.id),
            "skip_invalid": "false",
        },
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["success"] is False
    assert body["imported_count"] == 0
    assert len(body["transaction_ids"]) == 0
    assert len(body["errors"]) > 0

    # Ни одной транзакции не создано
    txs = (await session.execute(select(StockTransaction))).scalars().all()
    assert len(txs) == 0


async def test_import_remainders_excel_clear_existing(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """clear_existing=True: существующие остатки обнуляются, потом импортируются новые."""
    product = await _make_product(session, "CLEAR-001")
    location = await _make_location(session, "CLEAR-LOC")
    await session.commit()

    # Создаём существующий остаток 200 шт
    await _make_stock_balance(session, product.id, location.id, Decimal("200"))

    # Проверяем что баланс есть
    bal_before = await session.execute(
        select(StockBalance).where(
            StockBalance.product_id == product.id,
            StockBalance.location_id == location.id,
        )
    )
    assert bal_before.scalar_one_or_none() is not None

    excel_buf = _make_excel([
        ("CLEAR-001", 50, "новый импорт"),
    ])

    resp = await client.post(
        "/api/stock/import/remainders",
        files={"file": ("test.xlsx", excel_buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={
            "location_id": str(location.id),
            "clear_existing": "true",
        },
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["success"] is True
    assert body["imported_count"] == 1

    # Всего 3 транзакции: 1 (initial _make_stock_balance) +
    # 1 ADJUSTMENT_OUT (clear_existing) + 1 MANUAL_IN (import).
    txs = (await session.execute(select(StockTransaction))).scalars().all()
    assert len(txs) == 3

    reasons = [tx.reason for tx in txs]
    assert Reason.ADJUSTMENT_OUT in reasons
    assert Reason.MANUAL_IN in reasons

    adjust_tx = next(tx for tx in txs if tx.reason == Reason.ADJUSTMENT_OUT)
    assert float(adjust_tx.quantity) == 200.0
    assert adjust_tx.from_location_id == location.id

    manual_txs = [tx for tx in txs if tx.reason == Reason.MANUAL_IN]
    assert len(manual_txs) == 2  # 1 from _make_stock_balance + 1 from import
    import_manual = next(
        tx for tx in manual_txs
        if tx.source_ref and tx.source_ref.startswith("import_remainders:")
    )
    assert float(import_manual.quantity) == 50.0

    # Финальный баланс = 50
    final_balance = await session.execute(
        select(StockBalance).where(
            StockBalance.product_id == product.id,
            StockBalance.location_id == location.id,
        )
    )
    bal = final_balance.scalar_one_or_none()
    assert bal is not None
    assert float(bal.balance_qty) == 50.0


async def test_import_remainders_excel_quality_state(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """Колонка «Статус качества»=Брак → StockBalance с quality_state=SCRAP."""
    product = await _make_product(session, "SCRAP-001")
    location = await _make_location(session, "SCRAP-LOC")
    await session.commit()

    excel_buf = _make_excel(
        [("SCRAP-001", 15, "Брак")],
        headers=("SKU", "Количество", "Статус качества"),
    )

    resp = await client.post(
        "/api/stock/import/remainders",
        files={"file": ("test.xlsx", excel_buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={
            "location_id": str(location.id),
        },
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["success"] is True
    assert body["imported_count"] == 1

    # Проверяем транзакцию
    tx = (await session.execute(select(StockTransaction))).scalar_one()
    assert tx.to_quality_state == QualityState.SCRAP

    balance = await session.execute(
        select(StockBalance).where(
            StockBalance.product_id == product.id,
            StockBalance.location_id == location.id,
            StockBalance.quality_state == QualityState.SCRAP,
        )
    )
    bal = balance.scalar_one_or_none()
    assert bal is not None
    assert float(bal.balance_qty) == 15.0

    good_balance = await session.execute(
        select(StockBalance).where(
            StockBalance.product_id == product.id,
            StockBalance.location_id == location.id,
            StockBalance.quality_state == QualityState.GOOD,
        )
    )
    assert good_balance.scalar_one_or_none() is None


async def test_import_remainders_excel_per_row_quality_states(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """Разные статусы качества в одном файле импортируются построчно."""
    good_product = await _make_product(session, "MIX-GOOD")
    defect_product = await _make_product(session, "MIX-DEFECT")
    final_product = await _make_product(session, "MIX-FINAL")
    location = await _make_location(session, "MIX-LOC")
    await session.commit()

    excel_buf = _make_excel(
        [
            ("MIX-GOOD", 10, "Годный"),
            ("MIX-DEFECT", 5, "Брак"),
            ("MIX-FINAL", 3, "Окончательный брак"),
        ],
        headers=("SKU", "Количество", "Статус качества"),
    )

    resp = await client.post(
        "/api/stock/import/remainders",
        files={"file": ("test.xlsx", excel_buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"location_id": str(location.id)},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["success"] is True
    assert body["imported_count"] == 3

    txs = (await session.execute(select(StockTransaction))).scalars().all()
    assert len(txs) == 3
    by_product = {tx.product_id: tx for tx in txs}
    assert by_product[good_product.id].to_quality_state == QualityState.GOOD
    assert by_product[defect_product.id].to_quality_state == QualityState.SCRAP
    assert by_product[final_product.id].to_quality_state == QualityState.FINAL_SCRAP


async def test_preview_remainders_excel_rejects_rework_quality_state(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """Переделка не поддерживается при импорте остатков."""
    await _make_product(session, "RW-ERR")
    location = await _make_location(session, "RW-LOC")
    await session.commit()

    excel_buf = _make_excel(
        [("RW-ERR", 10, "Переделка")],
        headers=("SKU", "Количество", "Статус качества"),
    )

    resp = await client.post(
        "/api/stock/import/remainders/preview",
        files={"file": ("test.xlsx", excel_buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"location_id": str(location.id)},
    )
    assert resp.status_code == 200, resp.text
    item = resp.json()["items"][0]
    assert item["status"] == "invalid"
    assert any("переделка" in err.lower() for err in item["errors"])


async def test_preview_remainders_excel_unknown_quality_state(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """Неизвестный статус качества помечает строку invalid."""
    product = await _make_product(session, "QTY-ERR")
    location = await _make_location(session, "QTY-LOC")
    await session.commit()

    excel_buf = _make_excel(
        [("QTY-ERR", 10, "повреждённый")],
        headers=("SKU", "Количество", "Статус качества"),
    )

    resp = await client.post(
        "/api/stock/import/remainders/preview",
        files={"file": ("test.xlsx", excel_buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"location_id": str(location.id)},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    item = body["items"][0]
    assert item["status"] == "invalid"
    assert any("статус качества" in err.lower() for err in item["errors"])


async def test_import_remainders_excel_unknown_sku(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """SKU которого нет в БД → invalid."""
    location = await _make_location(session, "UNK-LOC")
    await session.commit()

    excel_buf = _make_excel([
        ("NONEXISTENT-SKU", 99, "нет такого продукта"),
    ])

    resp = await client.post(
        "/api/stock/import/remainders",
        files={"file": ("test.xlsx", excel_buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={
            "location_id": str(location.id),
            "skip_invalid": "true",
        },
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    # skip_invalid=True → success, но imported_count=0
    assert body["success"] is True
    assert body["imported_count"] == 0
    assert len(body["errors"]) == 1
    assert "not found" in body["errors"][0]


async def test_download_remainders_template(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """GET /import/remainders/template возвращает валидный .xlsx файл."""
    location = await _make_location(session, "TMPL-LOC")
    await session.commit()

    resp = await client.get(
        "/api/stock/import/remainders/template",
        params={"location_id": location.id},
    )
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "filename*" in resp.headers.get("content-disposition", "")

    # Проверяем что это валидный xlsx
    content = resp.content
    assert len(content) > 100
    # Проверяем, что можно открыть openpyxl
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(content))
    assert wb.active is not None
    ws = wb.active
    # заголовок + 3 примера = 4 строки
    assert ws.max_row == 4
    assert ws.cell(1, 1).value == "Артикул"
    assert ws.cell(1, 3).value == "Статус качества"
    assert ws.cell(1, 4).value == "Операции"
    assert ws.cell(1, 5).value == "Участок"
    assert ws.cell(2, 1).value == "361"
    assert ws.cell(3, 1).value == "ALS-1289"
    assert ws.cell(4, 1).value == "ЮП-2630"


async def test_download_remainders_template_example_rows_match_modal(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """Шаблон содержит те же примеры строк, что и модалка импорта."""
    location = await _make_location(session, "TMPL-OPS-LOC")
    await session.commit()

    resp = await client.get(
        "/api/stock/import/remainders/template",
        params={"location_id": location.id},
    )
    assert resp.status_code == 200, resp.text

    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active
    assert ws is not None

    assert ws.cell(3, 4).value == "Дробеструй"
    row3_ops = str(ws.cell(4, 4).value or "")
    assert "Дробеструй" in row3_ops
    assert "Чёрный" in row3_ops
    assert "Стрейч" in row3_ops
    assert ws.cell(4, 3).value == "Окончательный брак"
    assert ws.cell(4, 6).value == "Срочный заказ"


# ─── Tests: Preview with completed stages & target section ──────────────────


async def test_preview_returns_completed_stages(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """Preview возвращает completed_stages для строк с выполненными операциями,
    с дедупликацией повторяющихся имён операций (R2, edge-case)."""
    await _make_product(session, "CS-001")
    location = await _make_location(session, "CS-LOC")
    section = Section(
        code="PRESSING", name="Прессовый участок", type="production",
        is_active=True, sort_order=0,
    )
    session.add(section)
    await session.flush()
    await _make_section_operation(session, section, "Прессование", sort_order=10)
    await session.commit()

    excel_buf = _make_excel(
        [("CS-001", 100, "PRESSING", "Прессование, Прессование, Прессование", "коммент")],
        headers=("SKU", "Количество", "Целевая секция", "Выполненные операции", "Комментарий"),
    )
    resp = await client.post(
        "/api/stock/import/remainders/preview",
        files={"file": ("test.xlsx", excel_buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"location_id": str(location.id), "sheet_index": "0"},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert len(body["items"]) == 1
    item = body["items"][0]
    # dedup: "Прессование, Прессование, Прессование" → 1 элемент
    assert len(item["completed_stages"]) == 1
    stage = item["completed_stages"][0]
    assert stage["operation_name"] == "Прессование"
    assert stage["section_code"] == "PRESSING"
    assert stage["is_significant"] is True
    assert stage["sequence"] == section.sort_order


async def test_preview_completed_stages_use_section_order_not_intra_section(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """SHOT и ANOD не должны группироваться как «совмещено» из-за одинакового sort_order операции."""
    await _make_product(session, "ROUTE-001")
    location = await _make_location(session, "ROUTE-LOC")
    shot_section = Section(
        code="SHOT-TEST",
        name="Дробеструй",
        type="production",
        is_active=True,
        sort_order=40,
    )
    anod_section = Section(
        code="ANOD-TEST",
        name="Анодирование",
        type="production",
        is_active=True,
        sort_order=50,
    )
    session.add_all([shot_section, anod_section])
    await session.flush()
    await _make_section_operation(session, shot_section, "Дробеструй", operation_code="SHOT", sort_order=10)
    await _make_section_operation(session, anod_section, "Чёрный", operation_code="ANOD_05", sort_order=10)
    await _make_section_operation(
        session, anod_section, "Стрейч", operation_code="PACK_STRETCH", sort_order=20,
    )
    await session.commit()

    excel_buf = _make_excel(
        [("ROUTE-001", 100, "", "Дробеструй, Чёрный, Стрейч", "")],
        headers=("SKU", "Количество", "Целевая секция", "Выполненные операции", "Комментарий"),
    )
    resp = await client.post(
        "/api/stock/import/remainders/preview",
        files={"file": ("test.xlsx", excel_buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"location_id": str(location.id), "sheet_index": "0"},
    )
    assert resp.status_code == 200, resp.text
    stages = resp.json()["items"][0]["completed_stages"]
    assert len(stages) == 3
    assert stages[0]["operation_name"] == "Дробеструй"
    assert stages[0]["sequence"] == 40
    assert stages[0]["section_code"] == "SHOT-TEST"
    assert stages[1]["operation_name"] == "Чёрный"
    assert stages[1]["sequence"] == 50
    assert stages[2]["operation_name"] == "Стрейч"
    assert stages[2]["sequence"] == 50
    assert len({s["sequence"] for s in stages[:2]}) == 2


async def test_preview_target_section_name_resolved(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """Preview резолвит имя целевой секции в ID (R3)."""
    await _make_product(session, "TSR-001")
    location = await _make_location(session, "TSR-LOC")
    target_sec = await _make_warehouse_section(
        session, "WIP-A", "Цех WIP-A", type="wip_stock",
    )
    await session.commit()

    excel_buf = _make_excel(
        [("TSR-001", 100, "Цех WIP-A", "", "")],
        headers=("SKU", "Количество", "Целевая секция", "Выполненные операции", "Комментарий"),
    )
    resp = await client.post(
        "/api/stock/import/remainders/preview",
        files={"file": ("test.xlsx", excel_buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"location_id": str(location.id), "sheet_index": "0"},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert len(body["items"]) == 1
    item = body["items"][0]
    assert item["target_section_id"] == target_sec.id
    assert item["target_section_name"] == "Цех WIP-A"
    assert item["errors"] == []
    assert item["status"] == "valid"


async def test_preview_target_section_unknown_warns_but_not_fails(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """Неизвестная целевая секция → warning, но статус valid (R3)."""
    await _make_product(session, "UNK-SEC")
    location = await _make_location(session, "UNK-SEC-LOC")
    await session.commit()

    excel_buf = _make_excel(
        [("UNK-SEC", 100, "Несуществующая секция", "", "")],
        headers=("SKU", "Количество", "Целевая секция", "Выполненные операции", "Комментарий"),
    )
    resp = await client.post(
        "/api/stock/import/remainders/preview",
        files={"file": ("test.xlsx", excel_buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"location_id": str(location.id), "sheet_index": "0"},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert len(body["items"]) == 1
    item = body["items"][0]
    assert item["target_section_id"] is None
    assert item["target_section_name"] == "Несуществующая секция"
    assert any("не найден" in e for e in item["errors"])
    assert item["status"] == "valid"


async def test_preview_clipboard_text_parses_rows(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """Preview принимает TSV из буфера обмена вместо Excel-файла."""
    await _make_product(session, "CLIP-001")
    location = await _make_location(session, "CLIP-LOC")
    await session.commit()

    clipboard = _make_clipboard_tsv(
        [("CLIP-001", 42, "", "Дробеструй", "из буфера")],
        headers=("Артикул", "Количество", "Участок", "Выполненные операции", "Комментарий"),
    )
    resp = await client.post(
        "/api/stock/import/remainders/preview",
        data={
            "location_id": str(location.id),
            "clipboard_text": clipboard,
        },
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["sheet_name"] == "Буфер обмена"
    assert len(body["items"]) == 1
    item = body["items"][0]
    assert item["sku"] == "CLIP-001"
    assert item["quantity"] == 42
    assert item["comment"] == "из буфера"
    assert item["status"] == "valid"


async def test_import_clipboard_creates_balance(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """Импорт из буфера создаёт MANUAL_IN и обновляет баланс."""
    product = await _make_product(session, "CLIP-IMP")
    location = await _make_location(session, "CLIP-IMP-LOC")
    await session.commit()

    clipboard = _make_clipboard_tsv(
        [("CLIP-IMP", 15, "", "", "")],
        headers=("SKU", "Количество", "Комментарий"),
    )
    resp = await client.post(
        "/api/stock/import/remainders",
        data={
            "location_id": str(location.id),
            "clipboard_text": clipboard,
            "skip_invalid": "true",
        },
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["success"] is True
    assert body["imported_count"] == 1

    balance = await session.scalar(
        select(StockBalance).where(
            StockBalance.product_id == product.id,
            StockBalance.location_id == location.id,
            StockBalance.quality_state == QualityState.GOOD,
        )
    )
    assert balance is not None
    assert balance.balance_qty == Decimal("15")


async def test_preview_clipboard_two_columns_without_quantity(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """TSV с артикулом и операциями без колонки количества — qty=1 по умолчанию."""
    await _make_product(session, "ЮП-460")
    location = await _make_location(session, "CLIP-2COL")
    await session.commit()

    clipboard = _make_clipboard_tsv(
        [
            ("ЮП-460", "Окно, Дробеструй"),
            ("ЮП-460", "Гребенка"),
        ],
        headers=("Артикул", "Выполненные операции"),
    )
    resp = await client.post(
        "/api/stock/import/remainders/preview",
        data={"clipboard_text": clipboard},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["summary"]["total"] == 2
    assert body["summary"]["valid"] == 2
    for item in body["items"]:
        assert item["sku"] == "ЮП-460"
        assert item["quantity"] == 1


async def test_preview_positional_columns_without_headers(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """Без заголовков столбцы читаются в порядке шаблона Excel."""
    await _make_product(session, "POS-001")
    location = await _make_location(session, "POS-LOC")
    await session.commit()

    clipboard = (
        f"POS-001\t100\tГодный\tДробеструй\t{location.name}\tПартия A\n"
        f"POS-001\t50\tБрак\t\t{location.name}\t"
    )
    resp = await client.post(
        "/api/stock/import/remainders/preview",
        data={"clipboard_text": clipboard},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["summary"]["total"] == 2
    assert body["summary"]["valid"] == 2

    first, second = body["items"]
    assert first["sku"] == "POS-001"
    assert first["quantity"] == 100
    assert first["target_section_name"] == location.name
    assert first["quality_state"] == "good"
    assert first["completed_operations_raw"] == "Дробеструй"
    assert first["comment"] == "Партия A"

    assert second["quantity"] == 50
    assert second["quality_state"] == "scrap"


async def test_preview_positional_sparse_columns_default_qty(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """Без заголовков: только артикул — количество по умолчанию 1."""
    await _make_product(session, "POS-SPARSE")
    await session.commit()

    clipboard = "POS-SPARSE"
    resp = await client.post(
        "/api/stock/import/remainders/preview",
        data={"clipboard_text": clipboard},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["summary"]["total"] == 1
    assert body["summary"]["valid"] == 1
    assert body["items"][0]["sku"] == "POS-SPARSE"
    assert body["items"][0]["quantity"] == 1


async def test_preview_clipboard_concatenated_without_tabs(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """Склеенный paste без табуляции: Артикул + Выполненные операции."""
    await _make_product(session, "ЮП-460")
    await session.commit()

    clipboard = (
        "АртикулВыполненные операции\n"
        "ЮП-460Окно, Дробеструй\n"
        "ЮП-460Гребенка, Дробеструй\n"
        "ЮП-460Окно"
    )
    resp = await client.post(
        "/api/stock/import/remainders/preview",
        data={"clipboard_text": clipboard},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["summary"]["total"] == 3
    assert body["summary"]["valid"] == 3
    items = body["items"]
    assert items[0]["sku"] == "ЮП-460"
    assert items[0]["quantity"] == 1
    assert "Окно" in (items[0]["completed_operations_raw"] or "")
    assert items[1]["completed_operations_raw"] is not None
    assert "Гребенка" in items[1]["completed_operations_raw"]


async def test_preview_clipboard_unrecognized_rows_still_returned(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """Нераспознанные строки возвращаются как invalid, а не пропускаются."""
    clipboard = "Артикул\tКоличество\n\t\nмусор без структуры"
    resp = await client.post(
        "/api/stock/import/remainders/preview",
        data={"clipboard_text": clipboard},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["summary"]["total"] >= 1
    assert body["summary"]["invalid"] >= 1
    assert any(item["errors"] for item in body["items"])


async def test_preview_target_section_production_type_rejected(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """production-секция как target → warning, не invalid (R4)."""
    await _make_product(session, "PROD-TGT")
    location = await _make_location(session, "PROD-TGT-LOC")
    prod_sec = Section(
        code="PROD-SEC", name="Производственный участок", type="production",
        is_active=True, sort_order=0,
    )
    session.add(prod_sec)
    await session.commit()

    excel_buf = _make_excel(
        [("PROD-TGT", 100, "Производственный участок", "", "")],
        headers=("SKU", "Количество", "Целевая секция", "Выполненные операции", "Комментарий"),
    )
    resp = await client.post(
        "/api/stock/import/remainders/preview",
        files={"file": ("test.xlsx", excel_buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"location_id": str(location.id), "sheet_index": "0"},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert len(body["items"]) == 1
    item = body["items"][0]
    assert item["target_section_id"] is None
    assert any("production" in e.lower() for e in item["errors"])


# ─── Tests: Import with per-row target section ──────────────────────────────


async def test_import_per_row_target_creates_distinct_balances(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """Per-row target section → разные балансы на разных секциях + integrity (R5)."""
    product = await _make_product(session, "PER-ROW")
    # S1 = raw_stock (используется как default location формы)
    s1 = await _make_warehouse_section(
        session, "DEFAULT", "Default Stock", type="raw_stock",
    )
    # S2 = wip_stock (целевая секция для второй строки)
    s2 = await _make_warehouse_section(
        session, "WIP-B", "WIP-B", type="wip_stock",
    )
    await session.commit()

    excel_buf = _make_excel(
        [
            ("PER-ROW", 50, "", "", "row1"),
            ("PER-ROW", 30, "WIP-B", "", "row2"),
        ],
        headers=("SKU", "Количество", "Целевая секция", "Выполненные операции", "Комментарий"),
    )
    resp = await client.post(
        "/api/stock/import/remainders",
        files={"file": ("test.xlsx", excel_buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={
            "location_id": str(s1.id),
            "skip_invalid": "true",
            "clear_existing": "false",
        },
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["success"] is True
    assert body["imported_count"] == 2

    # Проверяем транзакции
    txs = (await session.execute(select(StockTransaction))).scalars().all()
    assert len(txs) == 2
    for tx in txs:
        assert tx.reason == Reason.MANUAL_IN
        assert tx.from_location_id is None

    # Разные to_location_id
    to_locs = {tx.to_location_id for tx in txs}
    assert to_locs == {s1.id, s2.id}

    # Проверяем балансы
    bals = (await session.execute(select(StockBalance))).scalars().all()
    assert len(bals) == 2
    bal_by_loc = {b.location_id: float(b.balance_qty) for b in bals}
    assert bal_by_loc[s1.id] == 50.0
    assert bal_by_loc[s2.id] == 30.0

    # Integrity invariants (S1-S6)
    await assert_no_invariants_violations(session, context="after-per-row-target")


async def test_import_clear_existing_with_target_section_override_returns_422(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """clear_existing=True + target_section_overrides → 422 (R6)."""
    location = await _make_location(session, "CLR-OVR-LOC")
    await session.commit()

    excel_buf = _make_excel(
        [("SOME-SKU", 10, "", "", "")],
        headers=("SKU", "Количество", "Целевая секция", "Выполненные операции", "Комментарий"),
    )
    resp = await client.post(
        "/api/stock/import/remainders",
        files={"file": ("test.xlsx", excel_buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={
            "location_id": str(location.id),
            "clear_existing": "true",
            "target_section_overrides": '{"2": "999"}',
        },
    )
    assert resp.status_code == 422, resp.text
    detail = resp.json().get("detail", "")
    assert "clear_existing" in detail.lower()


async def test_operations_endpoint_returns_production_significant(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """GET /import/remainders/operations возвращает только is_significant + production (R7)."""
    # Секция production для Op1, Op2
    prod_sec = Section(
        code="OPS-PROD", name="Производство", type="production",
        is_active=True, sort_order=0,
        icon="Drill", icon_color="#3B82F6",
    )
    # Секция storage (wip_stock) для Op3 с operation_type='transport'
    storage_sec = Section(
        code="OPS-STOR", name="Склад", type="wip_stock",
        is_active=True, sort_order=1,
    )
    session.add_all([prod_sec, storage_sec])
    await session.flush()

    # Op1: is_significant=True, production → ДОЛЖЕН быть в ответе
    await _make_section_operation(
        session, prod_sec, "Токарная", operation_code="TURN", sort_order=1,
        icon="Wrench", icon_color="#EF4444",
    )
    # Op2: is_significant=False, production → НЕ должен быть в ответе
    await _make_section_operation(
        session, prod_sec, "Фрезерная", operation_code="MILL",
        is_significant=False, sort_order=2,
    )
    # Op3: is_significant=True, transport → НЕ должен быть в ответе
    await _make_section_operation(
        session, storage_sec, "Перемещение", operation_code="MOVE",
        is_significant=True, operation_type="transport", sort_order=3,
    )
    await session.commit()

    resp = await client.get("/api/stock/import/remainders/operations")
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert isinstance(data, list)
    assert len(data) == 1
    op = data[0]
    assert op["operation_name"] == "Токарная"
    assert op["section_code"] == "OPS-PROD"
    assert op["is_significant"] is True
    assert op["section_icon"] == "Drill"
    assert op["section_icon_color"] == "#3B82F6"
    assert op["op_icon"] == "Wrench"
    assert op["op_icon_color"] == "#EF4444"


def test_parse_operations_from_comment_extracts_names() -> None:
    assert parse_operations_from_comment(None) == []
    assert parse_operations_from_comment("Импорт остатков из Excel") == []
    assert parse_operations_from_comment("Партия A | операции: Дробеструй, Чёрный") == [
        "Дробеструй",
        "Чёрный",
    ]


async def test_preview_remainders_with_custom_template_column_mapping(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """template_id: позиции колонок берутся из column_mapping шаблона (#15).

    Кастомный шаблон кладёт артикул в колонку C. Файл без строки заголовков
    → позиционный режим по порядку колонок шаблона, а не по дефолту 0..6.
    """
    await _make_product(session, "CTMP-001", "Продукт кастомного шаблона")
    template = ImportTemplate(
        name="Custom Layout",
        code="custom-layout",
        is_active=True,
        sort_order=0,
        column_mapping={
            "_config": {"length_required": True},
            "quantity": {"column": "A", "header": "Количество"},
            "comment": {"column": "B", "header": "Комментарий"},
            "sku": {"column": "C", "header": "Артикул"},
            "length": {"column": "D", "header": "Длина"},
        },
    )
    session.add(template)
    await session.flush()
    await session.commit()

    # Без заголовков: Кол-во | Комментарий | Артикул | Длина
    wb = Workbook()
    ws = wb.active
    ws.title = "Остатки"
    ws.append(["5", "note", "CTMP-001", "2,7"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = await client.post(
        "/api/stock/import/remainders/preview",
        files={"file": ("custom.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"template_id": str(template.id), "sheet_index": "0"},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()

    assert body["summary"]["valid"] == 1, body["items"]
    item = body["items"][0]
    assert item["sku"] == "CTMP-001"
    assert item["quantity"] == 5.0
    assert item["comment"] == "note"
    assert item["length_raw"] == "2,7"
    assert item["dimensions"] == {"length_mm": 2700}


async def test_preview_remainders_unknown_template_id_404(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """template_id несуществующего шаблона → 404."""
    excel_buf = _make_excel([("IMP-404", 1, "")])
    resp = await client.post(
        "/api/stock/import/remainders/preview",
        files={"file": ("test.xlsx", excel_buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"template_id": "999999", "sheet_index": "0"},
    )
    assert resp.status_code == 404, resp.text
