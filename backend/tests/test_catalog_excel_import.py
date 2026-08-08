"""Импорт справочника сырья из Excel (#63): preview-excel / apply-excel / template-excel."""

from io import BytesIO

import pytest
from httpx import AsyncClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.product import Product, ProductLength, ProductProcessingFlag, ProductType
from app.services.catalog_excel_import import TEMPLATE_HEADERS

PREVIEW_URL = "/api/catalog-import/preview-excel"
APPLY_URL = "/api/catalog-import/apply-excel"
TEMPLATE_URL = "/api/catalog-import/template-excel"

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _xlsx_bytes(rows: list[list], headers: list[str] | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(headers if headers is not None else list(TEMPLATE_HEADERS))
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _row(**kwargs) -> list:
    """Строка файла по именованным полям (остальные колонки пустые)."""
    values: dict[str, object] = {field: "" for field in (
        "sku", "name", "notes", "lengths", "perimeter", "mount_width",
        "quantities", "paired", "skip_shot", "laminated", "aliases",
    )}
    values.update(kwargs)
    return [
        values["sku"], values["name"], values["notes"], values["lengths"],
        values["perimeter"], values["mount_width"], values["quantities"],
        values["paired"], values["skip_shot"], values["laminated"], values["aliases"],
    ]


async def _make_product(
    session: AsyncSession,
    *,
    sku: str,
    name: str | None = None,
    lengths: list[float] | None = None,
    aliases: list[str] | None = None,
    **attrs,
) -> Product:
    product = Product(
        sku=sku,
        name=name or sku,
        type=ProductType.component,
        unit="шт",
        is_active=True,
        aliases=aliases or [],
    )
    for key, value in attrs.items():
        setattr(product, key, value)
    session.add(product)
    await session.flush()
    for length in lengths or []:
        session.add(ProductLength(product_id=product.id, length_mm=length))
    await session.flush()
    return product


async def _upload(client: AsyncClient, url: str, content: bytes, filename: str = "catalog.xlsx"):
    return await client.post(url, files={"file": (filename, content, XLSX_MIME)})


async def _product_lengths(session: AsyncSession, product_id: int) -> list[float]:
    rows = await session.scalars(
        select(ProductLength.length_mm).where(ProductLength.product_id == product_id)
    )
    return sorted(rows.all())


# ─── template-excel ──────────────────────────────────────────────────────────


async def test_template_excel_returns_full_reference_headers(client: AsyncClient) -> None:
    resp = await client.get(TEMPLATE_URL)
    assert resp.status_code == 200, resp.text
    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert headers == list(TEMPLATE_HEADERS)
    assert "Активен" not in headers and "Тип" not in headers


# ─── preview-excel ───────────────────────────────────────────────────────────


async def test_preview_excel_rejects_wrong_extension(client: AsyncClient) -> None:
    resp = await _upload(client, PREVIEW_URL, b"not a workbook", filename="catalog.txt")
    assert resp.status_code == 400


async def test_preview_excel_create_update_skip_actions(
    client: AsyncClient, session: AsyncSession
) -> None:
    await _make_product(
        session, sku="ЮП-100", name="Старое имя", perimeter_mm=50.0, lengths=[2780.0]
    )
    content = _xlsx_bytes([
        _row(sku="ЮП-100", perimeter="64,2"),  # update
        _row(sku="ЮП-200", name="Новый"),  # create
        _row(sku="ЮП-300"),  # create без данных
    ])
    resp = await _upload(client, PREVIEW_URL, content)
    assert resp.status_code == 200, resp.text
    body = resp.json()
    actions = {item["sku"]: item["action"] for item in body["items"]}
    assert actions == {"ЮП-100": "update", "ЮП-200": "create", "ЮП-300": "create"}
    assert body["stats"]["total"] == 3
    assert body["stats"]["create"] == 2
    assert body["stats"]["update"] == 1
    assert body["stats"]["skip"] == 0
    assert body["errors"] == []
    empty_create = next(item for item in body["items"] if item["sku"] == "ЮП-300")
    assert empty_create["warnings"]  # создание без данных — с warning


async def test_preview_excel_skip_when_nothing_changes(
    client: AsyncClient, session: AsyncSession
) -> None:
    await _make_product(session, sku="ЮП-100", perimeter_mm=64.2)
    content = _xlsx_bytes([_row(sku="ЮП-100", perimeter="64,2")])
    resp = await _upload(client, PREVIEW_URL, content)
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["items"][0]["action"] == "skip"
    assert body["stats"]["skip"] == 1


async def test_preview_excel_row_errors_reported_and_rows_skipped(
    client: AsyncClient, session: AsyncSession
) -> None:
    content = _xlsx_bytes([
        _row(sku="BAD-PERIM", perimeter="0"),
        _row(sku="BAD-PERIM-JUNK", perimeter="мусор"),
        _row(sku="BAD-MOUNT", mount_width="-5"),
        _row(sku="BAD-COUNT", lengths="2780, 3000", quantities="72"),
        _row(sku="BAD-QTY", lengths="2780", quantities="12,5"),
        _row(sku="BAD-BOOL", paired="может быть"),
        _row(sku="OK", perimeter="10"),
    ])
    resp = await _upload(client, PREVIEW_URL, content)
    assert resp.status_code == 200, resp.text
    body = resp.json()
    error_skus = {err["sku"] for err in body["errors"]}
    assert error_skus == {"BAD-PERIM", "BAD-PERIM-JUNK", "BAD-MOUNT", "BAD-COUNT", "BAD-QTY", "BAD-BOOL"}
    for err in body["errors"]:
        assert err["row"] >= 2
        assert err["message"]
    assert {item["sku"] for item in body["items"]} == {"OK"}


async def test_preview_excel_lookup_by_primary_sku_only(
    client: AsyncClient, session: AsyncSession
) -> None:
    await _make_product(session, sku="ЮП-100", aliases=["АЛИАС-1"])
    content = _xlsx_bytes([_row(sku="АЛИАС-1", perimeter="10")])
    resp = await _upload(client, PREVIEW_URL, content)
    assert resp.status_code == 200, resp.text
    body = resp.json()
    # Алиас в колонке «Артикул» → новый продукт, поиск по алиасам не идёт
    assert body["items"][0]["action"] == "create"


async def test_preview_excel_duplicate_sku_in_file_is_error(client: AsyncClient) -> None:
    content = _xlsx_bytes([
        _row(sku="ДУБЛЬ", perimeter="10"),
        _row(sku="ДУБЛЬ", perimeter="20"),
    ])
    resp = await _upload(client, PREVIEW_URL, content)
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert len(body["errors"]) == 1
    assert body["errors"][0]["sku"] == "ДУБЛЬ"
    assert len(body["items"]) == 1


# ─── apply-excel ─────────────────────────────────────────────────────────────


async def test_apply_excel_partial_update_touches_only_filled_cells(
    client: AsyncClient, session: AsyncSession
) -> None:
    product = await _make_product(
        session,
        sku="ЮП-100",
        name="Исходное имя",
        color="серебро",
        lengths=[2780.0],
        perimeter_mm=50.0,
    )
    product.quantity_per_hanger = {"2780": {"auto": None, "manual": 55}}
    await session.flush()

    content = _xlsx_bytes([
        _row(sku="ЮП-100", perimeter="64,2", lengths="2780, 3000", quantities="72, 65"),
    ])
    resp = await _upload(client, APPLY_URL, content)
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["updated"] == 1
    assert body["imported"] == 0
    assert body["errors"] == []

    await session.refresh(product)
    assert product.name == "Исходное имя"  # пустая ячейка не тронута
    assert product.color == "серебро"
    assert product.perimeter_mm == pytest.approx(64.2)
    assert await _product_lengths(session, product.id) == [2780.0, 3000.0]
    # quantity_per_hanger заменяется целиком, параллельно длинам
    assert product.quantity_per_hanger_by_length == {
        "2780": {"auto": None, "manual": 72},
        "3000": {"auto": None, "manual": 65},
    }


async def test_apply_excel_quantities_replace_whole_dict(
    client: AsyncClient, session: AsyncSession
) -> None:
    product = await _make_product(session, sku="ЮП-100", lengths=[2780.0, 3000.0])
    product.quantity_per_hanger = {
        "2780": {"auto": None, "manual": 55},
        "3000": {"auto": None, "manual": 50},
    }
    await session.flush()

    content = _xlsx_bytes([_row(sku="ЮП-100", lengths="3000", quantities="10")])
    resp = await _upload(client, APPLY_URL, content)
    assert resp.status_code == 200, resp.text
    await session.refresh(product)
    assert product.quantity_per_hanger_by_length == {"3000": {"auto": None, "manual": 10}}


async def test_apply_excel_quantities_for_existing_lengths_when_lengths_empty(
    client: AsyncClient, session: AsyncSession
) -> None:
    product = await _make_product(session, sku="ЮП-100", lengths=[2780.0, 3000.0])
    content = _xlsx_bytes([_row(sku="ЮП-100", quantities="72, 65")])
    resp = await _upload(client, APPLY_URL, content)
    assert resp.status_code == 200, resp.text
    await session.refresh(product)
    assert product.quantity_per_hanger_by_length == {
        "2780": {"auto": None, "manual": 72},
        "3000": {"auto": None, "manual": 65},
    }


async def test_apply_excel_creates_component_active(
    client: AsyncClient, session: AsyncSession
) -> None:
    content = _xlsx_bytes([
        _row(
            sku="ЮП-900",
            name="Профиль 900",
            lengths="2780, 3000",
            perimeter="64,2",
            mount_width="19,35",
            quantities="72, 65",
            paired="да",
            skip_shot="нет",
            laminated="да",
            aliases="ЭВ-1; ЭВ-2",
            notes="прим",
        ),
    ])
    resp = await _upload(client, APPLY_URL, content)
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["imported"] == 1
    assert body["errors"] == []

    product = await session.scalar(select(Product).where(Product.sku == "ЮП-900"))
    assert product is not None
    assert product.type == ProductType.component
    assert product.is_active is True
    assert product.name == "Профиль 900"
    assert product.is_paired_profile is True
    assert product.perimeter_mm == pytest.approx(64.2)
    assert product.mount_width_mm == pytest.approx(19.35)
    assert await _product_lengths(session, product.id) == [2780.0, 3000.0]
    assert product.quantity_per_hanger_by_length == {
        "2780": {"auto": None, "manual": 72},
        "3000": {"auto": None, "manual": 65},
    }
    assert sorted(product.aliases) == ["ЭВ-1", "ЭВ-2"]
    # Эквиваленты двунаправленные
    ev1 = await session.scalar(select(Product).where(Product.sku == "ЭВ-1"))
    assert ev1 is None  # неизвестные эквиваленты не создают продукты
    assert product.notes == "прим"


async def test_apply_excel_booleans_sync_flags(
    client: AsyncClient, session: AsyncSession
) -> None:
    from app.models.product import ProcessingFlag

    flag_shot = ProcessingFlag(code="skip_shot_blast", name="Не дробеструится")
    flag_lam = ProcessingFlag(code="is_laminated", name="Ламируется")
    session.add_all([flag_shot, flag_lam])
    await session.flush()

    product = await _make_product(session, sku="ЮП-100")
    session.add(ProductProcessingFlag(product_id=product.id, flag_id=flag_shot.id))
    await session.flush()

    content = _xlsx_bytes([_row(sku="ЮП-100", skip_shot="нет", laminated="да")])
    resp = await _upload(client, APPLY_URL, content)
    assert resp.status_code == 200, resp.text

    await session.refresh(product, attribute_names=["processing_flags"])
    codes = {f.code for f in product.processing_flags}
    assert codes == {"is_laminated"}


async def test_apply_excel_aliases_bidirectional(
    client: AsyncClient, session: AsyncSession
) -> None:
    await _make_product(session, sku="ЭВ-1")
    content = _xlsx_bytes([_row(sku="ЮП-100", aliases="ЭВ-1")])
    resp = await _upload(client, APPLY_URL, content)
    assert resp.status_code == 200, resp.text

    target = await session.scalar(select(Product).where(Product.sku == "ЭВ-1"))
    assert "ЮП-100" in (target.aliases or [])


async def test_apply_excel_counts_and_error_rows_skipped(
    client: AsyncClient, session: AsyncSession
) -> None:
    await _make_product(session, sku="ЮП-100", perimeter_mm=64.2)
    content = _xlsx_bytes([
        _row(sku="ЮП-100"),  # skip — ничего не меняется (кроме форсов type/active, которые уже верны)
        _row(sku="ЮП-200", perimeter="10"),  # imported
        _row(sku="ЮП-300", perimeter="-1"),  # ошибка строки
    ])
    resp = await _upload(client, APPLY_URL, content)
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["imported"] == 1
    assert body["skipped"] == 1
    assert body["updated"] == 0
    assert len(body["errors"]) == 1
    assert body["errors"][0]["sku"] == "ЮП-300"
    assert await session.scalar(select(Product).where(Product.sku == "ЮП-300")) is None


async def test_apply_excel_empty_row_creates_with_warning(
    client: AsyncClient, session: AsyncSession
) -> None:
    content = _xlsx_bytes([_row(sku="ЮП-ПУСТО")])
    preview = await _upload(client, PREVIEW_URL, content)
    assert preview.status_code == 200
    item = preview.json()["items"][0]
    assert item["action"] == "create"
    assert item["warnings"]

    resp = await _upload(client, APPLY_URL, content)
    assert resp.status_code == 200
    assert resp.json()["imported"] == 1
    product = await session.scalar(select(Product).where(Product.sku == "ЮП-ПУСТО"))
    assert product is not None
    assert product.type == ProductType.component
    assert product.is_active is True
