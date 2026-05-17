from io import BytesIO

import pytest
from openpyxl import Workbook

from app.core.config import settings
from app.models.import_template import ImportTemplate
from app.models.imports import ImportBatch, ImportFile
from app.models.production_plan import PlanChangeItem, PlanChangeSet, ProductionPlan
from app.models.route import RouteRuleProfile
from app.services.excel_import import parse_factory_plan_workbook, parse_row_selection
from app.services.import_normalization import default_import_normalization_rules


def _workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "План май 26 05"
    ws.append(["", "", "Комментарий"])
    ws.append(["Заявка № 05", "май"])
    ws.append([])
    ws.append(["", "", "", "", "", "", "", "", "", "", "", "", "Формирование ящиков"])
    ws.append(
        [
            "Артикул",
            "пополнение",
            "Наименование",
            "остатки сырья на КТМ",
            "Цвет",
            "кол-во шт. в 2,7",
            "Длина, м",
            "Пробивка/сверловка",
            "Упаковка",
            "Примечание ",
            "Длина после упак, м",
            "кол-во штук готовой продукции",
            "Запад",
            "Восток",
            "Вид конечного продукта",
            "Комментарии",
            "",
            "",
            "Упаковка в 1,8",
            "добавить",
        ]
    )
    ws.append(
        [
            "ЮП-2616",
            "ТЗ",
            "Кант универсальный 47мм 2,7 анод черный мат",
            7300,
            "черный",
            300,
            2.7,
            "",
            "смотка спанбондом поштучно в пачке 10 штук",
            "",
            2.7,
            300,
            "",
            300,
            "П/ф",
        ]
    )
    ws.append(["ЮП-2604", "ТЗ", "", 3700, "черный", 300, 2.7, "", "", "", 2.7, 300, "", 300, "П/ф"])
    ws.append(
        [
            "ЮП-2083",
            "ТЗ",
            "Стык 38 мм. 2,7 анод.серебро, матовый",
            2958,
            "серебро",
            1100,
            2.7,
            "сверло",
            "поф",
            "",
            0.9,
            1500,
            1500,
            0,
            "ГП",
        ]
    )
    ws.append(["ЮП-2083", "", "", "", "серебро", "", "", "сверло", "поф", "", 1.8, 900, 900, 0, "ГП"])
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


async def _create_template(session, *, name: str, code: str) -> ImportTemplate:
    template = ImportTemplate(
        name=name,
        code=code,
        is_active=True,
        column_mapping={"sku": {"header": "Артикул", "column": "A"}},
        normalization_rules=default_import_normalization_rules(),
    )
    session.add(template)
    await session.flush()
    return template


def test_factory_plan_parser_groups_paired_profiles_and_continuations() -> None:
    parsed = parse_factory_plan_workbook(_workbook_bytes(), "plan.xlsx")

    assert parsed.sheet_name == "План май 26 05"
    assert parsed.header_row_number == 5
    assert parsed.period_start.isoformat() == "2026-05-01"
    assert parsed.period_end.isoformat() == "2026-05-31"
    assert len(parsed.parsed_rows) == 3

    paired = parsed.parsed_rows[0]
    assert paired.source_row_numbers == [6, 7]
    assert paired.source_sku == "ЮП-2616+ЮП-2604"
    assert paired.quantity == 300
    assert paired.payload["paired_profile"] is True
    assert [component["sku"] for component in paired.payload["components"]] == ["ЮП-2616", "ЮП-2604"]
    assert paired.payload["raw_columns_meta"][0]["index"] == 1
    assert paired.payload["raw_columns_meta"][0]["letter"] == "A"
    assert paired.payload["raw_columns_meta"][0]["header"] == "Артикул"
    assert paired.payload["raw_columns_meta"][7]["index"] == 8
    assert paired.payload["raw_columns_meta"][7]["letter"] == "H"
    assert paired.payload["raw_columns_meta"][7]["header"] == "Пробивка/сверловка"

    continuation = parsed.parsed_rows[2]
    assert continuation.source_row_numbers == [9]
    assert continuation.payload["context_inherited"] is True
    assert continuation.source_name == "Стык 38 мм. 2,7 анод.серебро, матовый"
    assert continuation.quantity == 900


def test_parse_row_selection_csv_and_ranges() -> None:
    assert parse_row_selection("5") == {5}
    assert parse_row_selection("5,7,9") == {5, 7, 9}
    assert parse_row_selection("5-8") == {5, 6, 7, 8}
    assert parse_row_selection("5,7-9") == {5, 7, 8, 9}
    assert parse_row_selection("5, 7, 7, 9-10") == {5, 7, 9, 10}


@pytest.mark.parametrize("value", ["", "7-", "a", "15-12", "5,,7", "-1", "0"])
def test_parse_row_selection_invalid(value: str) -> None:
    with pytest.raises(ValueError):
        parse_row_selection(value)


def test_factory_plan_parser_row_selection_auto_includes_pair() -> None:
    parsed = parse_factory_plan_workbook(_workbook_bytes(), "plan.xlsx", row_selection="6")
    assert len(parsed.parsed_rows) == 1
    row = parsed.parsed_rows[0]
    assert row.source_row_numbers == [6, 7]
    assert any(w.startswith("paired_row_auto_included:") for w in row.warnings)
    assert parsed.selected_row_numbers == [6]
    assert parsed.auto_included_row_numbers == [7]


@pytest.mark.asyncio
async def test_import_excel_creates_batch_and_change_set(client, session, tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(settings, "IMPORT_STORAGE_DIR", str(tmp_path))
    template = await _create_template(session, name="Base Template", code="base-template")
    await session.commit()

    response = await client.post(
        f"/api/imports/excel?template_id={template.id}",
        files={
            "file": (
                "plan.xlsx",
                _workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201
    body = response.json()
    assert body["summary"]["total_positions"] == 3
    assert body["summary"]["paired_profile_positions"] == 1
    assert len(body["items"]) == 3
    assert body["items"][0]["after_data"]["source_sku"] == "ЮП-2616+ЮП-2604"
    assert body["items"][0]["after_data"]["operation_family"] == "NONE"
    assert body["items"][0]["after_data"]["output_kind"] == "semi_finished_shipment"
    assert body["items"][0]["after_data"]["has_pack_ops"] is False
    assert body["items"][0]["warnings"] == ["paired_profile_product_unmapped"]
    # ЮП-2083 not seeded in tests, so product_not_found is expected
    assert "product_not_found" in body["items"][1]["errors"] or "active_techcard_has_no_lines" in body["items"][1]["errors"]

    assert await session.get(ImportFile, body["import_file_id"]) is not None
    assert await session.get(ImportBatch, body["import_batch_id"]) is not None
    assert await session.get(ProductionPlan, body["production_plan_id"]) is not None
    assert await session.get(PlanChangeSet, body["change_set_id"]) is not None

    change_items = body["items"]
    assert await session.get(PlanChangeItem, change_items[0]["id"]) is not None


@pytest.mark.asyncio
async def test_preview_excel_resolves_paired_profile_warning_when_pair_techcard_exists(
    client, session, tmp_path, monkeypatch
) -> None:
    monkeypatch.setattr(settings, "IMPORT_STORAGE_DIR", str(tmp_path))

    from app.models.product import Product, ProductType
    from app.models.techcard import Techcard, TechcardLine

    template = await _create_template(session, name="Preview Pair Template", code="preview-pair-template")

    comp_a = Product(sku="ЮП-2616", name="Comp A", type=ProductType.component, unit="pcs", is_active=True)
    comp_b = Product(sku="ЮП-2604", name="Comp B", type=ProductType.component, unit="pcs", is_active=True)
    session.add_all([comp_a, comp_b])
    await session.flush()

    paired = Techcard(product_id=None, version="v1", is_active=True, processing_type="paired_processing")
    session.add(paired)
    await session.flush()
    session.add_all(
        [
            TechcardLine(techcard_id=paired.id, component_product_id=comp_a.id, quantity=1, unit="pcs"),
            TechcardLine(techcard_id=paired.id, component_product_id=comp_b.id, quantity=1, unit="pcs"),
        ]
    )
    await session.commit()

    response = await client.post(
        f"/api/imports/excel/preview?template_id={template.id}",
        files={
            "file": (
                "plan.xlsx",
                _workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    body = response.json()
    paired_item = body["items"][0]
    assert paired_item["source_sku"] == "ЮП-2616+ЮП-2604"
    assert "paired_profile_product_unmapped" not in paired_item["warnings"]


@pytest.mark.asyncio
async def test_import_excel_with_row_selection_filters_rows_and_reports_pair_autoinclude(
    client, session, tmp_path, monkeypatch
) -> None:
    monkeypatch.setattr(settings, "IMPORT_STORAGE_DIR", str(tmp_path))
    template = await _create_template(session, name="Rows Template", code="rows-template")
    await session.commit()

    response = await client.post(
        f"/api/imports/excel?template_id={template.id}",
        data={"row_selection": "6"},
        files={
            "file": (
                "plan.xlsx",
                _workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201
    body = response.json()
    assert body["summary"]["total_positions"] == 1
    assert body["summary"]["row_selection"] == "6"
    assert body["summary"]["selected_row_numbers"] == [6]
    assert body["summary"]["auto_included_row_numbers"] == [7]
    assert body["items"][0]["after_data"]["source_sku"] == "ЮП-2616+ЮП-2604"
    assert any(w.startswith("paired_row_auto_included:") for w in body["items"][0]["warnings"])


@pytest.mark.asyncio
async def test_import_excel_with_invalid_row_selection_returns_400(client, session, tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(settings, "IMPORT_STORAGE_DIR", str(tmp_path))
    template = await _create_template(session, name="Invalid Rows Template", code="invalid-rows-template")
    await session.commit()

    response = await client.post(
        f"/api/imports/excel?template_id={template.id}",
        data={"row_selection": "15-12"},
        files={
            "file": (
                "plan.xlsx",
                _workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 400
    assert "row range" in response.json()["detail"].lower()


@pytest.mark.asyncio
async def test_import_excel_requires_template_id(client, tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(settings, "IMPORT_STORAGE_DIR", str(tmp_path))

    response = await client.post(
        "/api/imports/excel",
        files={
            "file": (
                "plan.xlsx",
                _workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "template_id is required"


def test_factory_plan_parser_with_custom_mapping() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Custom Plan"
    ws.append(["SKU", "Name", "Qty", "Deadline", "Client", "Priority", "Order"])
    ws.append(["ABC-123", "Test Product", 50, "2026-06-15", "ClientA", 1, "ORD-001"])
    out = BytesIO()
    wb.save(out)
    content = out.getvalue()

    custom_mapping = {
        "sku": "SKU",
        "product_name": "Name",
        "quantity": "Qty",
        "due_date": "Deadline",
        "customer": "Client",
        "priority": "Priority",
        "order_ref": "Order",
    }
    parsed = parse_factory_plan_workbook(content, "custom.xlsx", column_mapping=custom_mapping)

    assert len(parsed.parsed_rows) == 1
    row = parsed.parsed_rows[0]
    assert row.source_sku == "ABC-123"
    assert row.source_name == "Test Product"
    assert row.quantity == 50
    assert row.payload["due_date"] == "2026-06-15"
    assert row.payload["customer"] == "ClientA"
    assert row.payload["priority"] == 1
    assert row.payload["order_ref"] == "ORD-001"


def test_factory_plan_parser_maps_additional_pack_operations() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "План май 26 05"
    ws.append(["", "", "Комментарий"])
    ws.append(["Заявка № 05", "май"])
    ws.append([])
    ws.append(["", "", "", "", "", "", "", "", "", "", "", "", "Формирование ящиков"])
    ws.append(
        [
            "Артикул",
            "пополнение",
            "Наименование",
            "остатки сырья на КТМ",
            "Цвет",
            "кол-во шт. в 2,7",
            "Длина, м",
            "Пробивка/сверловка",
            "Упаковка",
            "Примечание ",
            "Длина после упак, м",
            "кол-во штук готовой продукции",
            "Запад",
            "Восток",
            "Вид конечного продукта",
            "Комментарии",
        ]
    )
    ws.append(["SKU-GLUE", "", "Glue Profile", 0, "", 100, 2.7, "клей", "поф", "", 2.7, 100, "", 100, "ГП", ""])
    ws.append(["SKU-DIFF", "", "Diffuser Profile", 0, "", 200, 2.7, "рассеиватель", "поф", "", 2.7, 200, "", 200, "ГП", ""])
    ws.append(["SKU-NODIFF", "", "No Diffuser Profile", 0, "", 50, 2.7, "Без рассеивателя", "поф", "", 2.7, 50, "", 50, "П/ф", ""])
    out = BytesIO()
    wb.save(out)

    parsed = parse_factory_plan_workbook(
        out.getvalue(),
        "pack_ops.xlsx",
        normalization_rules=default_import_normalization_rules(),
    )
    assert len(parsed.parsed_rows) == 3

    glue = parsed.parsed_rows[0].payload
    assert glue["operation_code"] == "PACK"
    assert glue["operation_name"] == "Упаковка"
    assert glue["additional_pack_operations"][0]["operation_code"] == "PACK_GLUE"

    diffuser = parsed.parsed_rows[1].payload
    assert diffuser["operation_code"] == "PACK"
    assert diffuser["operation_name"] == "Упаковка"
    assert diffuser["additional_pack_operations"][0]["operation_code"] == "PACK_DIFFUSER"

    no_diff = parsed.parsed_rows[2].payload
    assert no_diff["operation_code"] == "PACK"
    assert no_diff["operation_name"] == "Упаковка"
    assert no_diff["additional_pack_operations"][0]["operation_code"] == "PACK_CUSTOM"


from datetime import date, datetime

from app.services.excel_import import _excel_date_to_date, _parse_date


def test_date_normalization() -> None:
    assert _parse_date(date(2026, 5, 2)) == date(2026, 5, 2)
    assert _parse_date(datetime(2026, 5, 2, 14, 30)) == date(2026, 5, 2)
    assert _parse_date("2026-05-02") == date(2026, 5, 2)
    assert _parse_date("02.05.2026") == date(2026, 5, 2)
    assert _excel_date_to_date(1) == date(1899, 12, 31)
    assert _parse_date(44561) == date(2021, 12, 31)
    assert _parse_date("invalid") is None
    assert _parse_date(None) is None


@pytest.mark.asyncio
async def test_replace_draft_mode_creates_cancel_for_missing_rows(client, session, tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(settings, "IMPORT_STORAGE_DIR", str(tmp_path))

    from app.models.techcard import Techcard, TechcardLine
    from app.models.product import Product, ProductType
    from app.models.route import ProductionRoute, RouteSignatureRule, RouteStep
    from app.models.routing import RouteOperationFamily, RouteOutputKind
    from app.models.section import Section

    product = Product(sku="FG-TEST", name="Test Product", type=ProductType.finished_good, unit="pcs")
    component = Product(sku="FG-TEST-RAW", name="Test Raw", type=ProductType.component, unit="pcs")
    sections = [
        Section(code="CUT", name="Cut"),
        Section(code="PACK", name="Pack"),
    ]
    session.add_all([product, component, *sections])
    await session.flush()

    techcard = Techcard(product_id=product.id, version="v1", is_active=True)
    session.add(techcard)
    await session.flush()
    session.add(TechcardLine(techcard_id=techcard.id, component_product_id=component.id, quantity=1, unit="pcs"))

    route = ProductionRoute(name="Main", is_active=True)
    session.add(route)
    await session.flush()
    session.add(
        RouteSignatureRule(
            route_id=route.id,
            operation_family=RouteOperationFamily.NONE,
            output_kind=RouteOutputKind.finished_good,
            has_pack_ops=False,
            priority=10,
            is_active=True,
        )
    )
    for index, section in enumerate(sections, start=1):
        session.add(
            RouteStep(
                route_id=route.id,
                sequence=index,
                section_id=section.id,
                operation_name=f"Step {index}",
                is_final=index == len(sections),
            )
        )
    await session.commit()

    def _make_workbook(rows: list[tuple[str, str, int]]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "План май 26 05"
        ws.append(["", "", "Комментарий"])
        ws.append(["Заявка № 05", "май"])
        ws.append([])
        ws.append(["", "", "", "", "", "", "", "", "", "", "", "", "Формирование ящиков"])
        ws.append(
            [
                "Артикул",
                "пополнение",
                "Наименование",
                "остатки сырья на КТМ",
                "Цвет",
                "кол-во шт. в 2,7",
                "Длина, м",
                "Пробивка/сверловка",
                "Упаковка",
                "Примечание ",
                "Длина после упак, м",
                "кол-во штук готовой продукции",
                "Запад",
                "Восток",
                "Вид конечного продукта",
                "Комментарии",
                "",
                "",
                "Упаковка в 1,8",
                "добавить",
            ]
        )
        for sku, name, qty in rows:
            ws.append([sku, "ТЗ", name, 0, "", qty, 2.7, "", "", "", 2.7, qty, "", qty, "ГП"])
        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    # First import with 2 rows
    template = await _create_template(session, name="Replace Template", code="replace-template")
    await session.commit()
    wb1 = _make_workbook([("FG-TEST", "Test Product", 100), ("FG-TEST", "Test Product", 200)])
    response1 = await client.post(
        f"/api/imports/excel?template_id={template.id}",
        files={"file": ("plan1.xlsx", wb1, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response1.status_code == 201
    body1 = response1.json()
    plan_id = body1["production_plan_id"]
    change_set_id1 = body1["change_set_id"]

    apply1 = await client.post(f"/api/production-plans/{plan_id}/change-sets/{change_set_id1}/apply")
    assert apply1.status_code == 200
    assert apply1.json()["created_positions"] == 2

    # Second import with 1 row (replace mode)
    wb2 = _make_workbook([("FG-TEST", "Test Product", 100)])
    response2 = await client.post(
        f"/api/imports/excel?template_id={template.id}",
        data={"mode": "replace_draft_from_same_source", "production_plan_id": str(plan_id)},
        files={"file": ("plan2.xlsx", wb2, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response2.status_code == 201
    body2 = response2.json()
    actions = [item["change_action"] for item in body2["items"]]
    assert "cancel_draft_position" in actions


@pytest.mark.asyncio
async def test_import_excel_resolves_profile_by_template_priority(client, session, tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(settings, "IMPORT_STORAGE_DIR", str(tmp_path))

    template = ImportTemplate(
        name="Template with Profiles",
        code="template-profiles",
        is_active=True,
        column_mapping={"sku": {"header": "Артикул", "column": "A"}},
        normalization_rules=default_import_normalization_rules(),
    )
    session.add(template)
    await session.flush()

    low_active = RouteRuleProfile(
        code="profile-low-active",
        name="Profile low active",
        is_active=True,
        priority=10,
        import_template_id=template.id,
    )
    high_inactive = RouteRuleProfile(
        code="profile-high-inactive",
        name="Profile high inactive",
        is_active=False,
        priority=999,
        import_template_id=template.id,
    )
    high_active = RouteRuleProfile(
        code="profile-high-active",
        name="Profile high active",
        is_active=True,
        priority=20,
        import_template_id=template.id,
    )
    session.add_all([low_active, high_inactive, high_active])
    await session.commit()

    response = await client.post(
        f"/api/imports/excel?template_id={template.id}",
        files={
            "file": (
                "plan.xlsx",
                _workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201
    body = response.json()
    assert body["template_id"] == template.id
    assert body["rule_profile_id"] == high_active.id


@pytest.mark.asyncio
async def test_import_excel_template_without_profile_uses_fallback(client, session, tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(settings, "IMPORT_STORAGE_DIR", str(tmp_path))

    template = ImportTemplate(
        name="Template without profile",
        code="template-no-profile",
        is_active=True,
        column_mapping={"sku": {"header": "Артикул", "column": "A"}},
        normalization_rules=default_import_normalization_rules(),
    )
    session.add(template)
    await session.commit()

    response = await client.post(
        f"/api/imports/excel?template_id={template.id}",
        files={
            "file": (
                "plan.xlsx",
                _workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201
    body = response.json()
    assert body["template_id"] == template.id
    assert body["rule_profile_id"] is None


@pytest.mark.asyncio
async def test_preview_excel_uses_template_profile_for_rule_selection(client, session, monkeypatch) -> None:
    template = ImportTemplate(
        name="Template preview profile",
        code="template-preview-profile",
        is_active=True,
        column_mapping={"sku": {"header": "Артикул", "column": "A"}},
        normalization_rules=default_import_normalization_rules(),
    )
    session.add(template)
    await session.flush()

    low_active = RouteRuleProfile(
        code="preview-profile-low",
        name="Preview profile low",
        is_active=True,
        priority=10,
        import_template_id=template.id,
    )
    high_inactive = RouteRuleProfile(
        code="preview-profile-high-inactive",
        name="Preview profile high inactive",
        is_active=False,
        priority=999,
        import_template_id=template.id,
    )
    high_active = RouteRuleProfile(
        code="preview-profile-high-active",
        name="Preview profile high active",
        is_active=True,
        priority=20,
        import_template_id=template.id,
    )
    session.add_all([low_active, high_inactive, high_active])
    await session.commit()

    captured: dict[str, int | None] = {"rule_profile_id": None}

    async def _fake_preview_excel_sheet(_db, **kwargs):
        captured["rule_profile_id"] = kwargs.get("rule_profile_id")
        return {
            "sheet_name": "test",
            "header_row_number": 1,
            "total_rows": 0,
            "summary": {},
            "items": [],
        }

    monkeypatch.setattr("app.services.plan_import_service.preview_excel_sheet", _fake_preview_excel_sheet)

    response = await client.post(
        f"/api/imports/excel/preview?template_id={template.id}",
        files={
            "file": (
                "plan.xlsx",
                _workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    assert captured["rule_profile_id"] == high_active.id
