"""Тесты импорта остатков с длиной (issue #5, ADR-0003 п. 3).

Проверяют:
- Парсинг колонки «Длина» (метры с запятой/точкой → мм) в Excel и позиционном режиме.
- Мусор/ноль в длине → invalid строка с ошибкой.
- Три ветки валидации: явная длина / типовой размер продукта / invalid.
- Продукт без обязательных измерений (крепёж) → dimensions = None.
- Применение импорта пишет dimensions в StockTransaction, баланс разбит по длинам.
- Шаблон импорта остатков содержит колонку «Длина».
"""
from __future__ import annotations

from io import BytesIO

import pytest
from httpx import AsyncClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Product, ProductType, Section
from app.models.dimension import DimensionType, ProductDimension
from app.stock.import_service import (
    generate_remainders_template_for_location,
    parse_remainders_excel,
)
from app.stock.models import StockBalance, StockTransaction
from tests.test_integrity_invariants import assert_no_invariants_violations

pytestmark = pytest.mark.asyncio

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ─── Helpers ───────────────────────────────────────────────────────────────────


async def _make_product(
    session: AsyncSession,
    sku: str = "LEN-PROD",
    name: str | None = None,
) -> Product:
    product = Product(
        sku=sku,
        name=name or sku,
        type=ProductType.finished_good,
        unit="pcs",
        is_active=True,
    )
    session.add(product)
    await session.flush()
    return product


async def _make_location(session: AsyncSession, code: str = "STOCK-LEN") -> Section:
    section = Section(
        code=code,
        name=code,
        type="raw_stock",
        is_active=True,
        sort_order=0,
    )
    session.add(section)
    await session.flush()
    return section


async def _link_length_dimension(
    session: AsyncSession,
    product: Product,
    *,
    is_required: bool = True,
    default_value: float | None = None,
) -> None:
    """Привязать измерение length_mm к продукту (создав тип при необходимости)."""
    dim_type = (
        await session.scalars(
            select(DimensionType).where(DimensionType.code == "length_mm")
        )
    ).first()
    if dim_type is None:
        dim_type = DimensionType(code="length_mm", name="Длина", unit="мм")
        session.add(dim_type)
        await session.flush()
    session.add(
        ProductDimension(
            product_id=product.id,
            dimension_type_id=dim_type.id,
            is_required=is_required,
            default_value=default_value,
        )
    )
    await session.flush()


def _make_excel(
    rows: list[tuple],
    headers: tuple[str, ...] | None = ("SKU", "Количество", "Длина", "Комментарий"),
) -> BytesIO:
    """Create an .xlsx file in memory; headers=None → позиционный режим."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Остатки"
    if headers is not None:
        ws.append(list(headers))
    for row in rows:
        ws.append(list(row))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── Tests: Parsing ────────────────────────────────────────────────────────────


async def test_parse_length_column_comma_and_dot() -> None:
    """«2,7» и «2.75» в колонке «Длина» → 2700 и 2750 мм."""
    excel_buf = _make_excel([
        ("LEN-001", 10, "2,7", None),
        ("LEN-001", 5, "2.75", None),
    ])
    _sheet, _total, items, summary = await parse_remainders_excel(excel_buf.getvalue())

    assert summary.total == 2
    assert summary.invalid == 0
    assert items[0].length_raw == "2,7"
    assert items[0].dimensions == {"length_mm": 2700}
    assert items[0].dimensions_label == "2,7 м"
    assert items[1].dimensions == {"length_mm": 2750}
    assert items[1].dimensions_label == "2,75 м"


async def test_parse_length_column_empty_gives_none() -> None:
    """Пустая ячейка и прочерк «—» → dimensions=None, строка валидна."""
    excel_buf = _make_excel([
        ("LEN-002", 10, "", None),
        ("LEN-002", 4, "—", None),
    ])
    _sheet, _total, items, _summary = await parse_remainders_excel(excel_buf.getvalue())

    assert all(it.status == "valid" for it in items)
    assert items[0].dimensions is None
    assert items[0].dimensions_label == "—"
    assert items[1].dimensions is None


async def test_parse_length_column_garbage_marks_invalid() -> None:
    """Мусор («abc») и ноль в длине → invalid с понятной ошибкой."""
    excel_buf = _make_excel([
        ("LEN-003", 10, "abc", None),
        ("LEN-003", 5, "0", None),
    ])
    _sheet, _total, items, summary = await parse_remainders_excel(excel_buf.getvalue())

    assert summary.invalid == 2
    assert items[0].status == "invalid"
    assert any("распознать длину" in e for e in items[0].errors)
    assert items[0].dimensions is None
    assert items[1].status == "invalid"
    assert any("положительной" in e for e in items[1].errors)


async def test_parse_length_positional_template_order() -> None:
    """Без заголовков «Длина» читается из 7-й колонки (порядок шаблона)."""
    excel_buf = _make_excel(
        [("LEN-POS", 12, "Годный", "", "", "", "1,35")],
        headers=None,
    )
    _sheet, _total, items, _summary = await parse_remainders_excel(excel_buf.getvalue())

    assert len(items) == 1
    assert items[0].sku == "LEN-POS"
    assert items[0].dimensions == {"length_mm": 1350}


async def test_parse_length_positional_old_six_columns_backcompat() -> None:
    """Старый шестиколоночный файл без заголовков парсится как раньше (без длины)."""
    excel_buf = _make_excel(
        [("LEN-OLD", 7, "Годный", "", "", "коммент")],
        headers=None,
    )
    _sheet, _total, items, summary = await parse_remainders_excel(excel_buf.getvalue())

    assert summary.invalid == 0
    assert items[0].dimensions is None
    assert items[0].length_raw is None


# ─── Tests: Validation branches (preview) ─────────────────────────────────────


async def test_preview_explicit_length_wins_over_default(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """Ветка 1: колонка заполнена → берём её, а не типовой размер."""
    product = await _make_product(session, "LEN-EXPL")
    await _link_length_dimension(session, product, is_required=True, default_value=2700)
    await session.commit()

    excel_buf = _make_excel([("LEN-EXPL", 10, "0,9", None)])
    resp = await client.post(
        "/api/stock/import/remainders/preview",
        files={"file": ("test.xlsx", excel_buf, XLSX_MIME)},
        data={"sheet_index": "0"},
    )
    assert resp.status_code == 200, resp.text
    item = resp.json()["items"][0]
    assert item["status"] == "valid"
    assert item["dimensions"] == {"length_mm": 900}
    assert item["dimensions_label"] == "0,9 м"


async def test_preview_empty_length_falls_back_to_default(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """Ветка 2: пустая колонка → типовой размер продукта из справочника."""
    product = await _make_product(session, "LEN-DEF")
    await _link_length_dimension(session, product, is_required=True, default_value=2700)
    await session.commit()

    excel_buf = _make_excel([("LEN-DEF", 10, "", None)])
    resp = await client.post(
        "/api/stock/import/remainders/preview",
        files={"file": ("test.xlsx", excel_buf, XLSX_MIME)},
        data={"sheet_index": "0"},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["summary"]["invalid"] == 0
    item = body["items"][0]
    assert item["status"] == "valid"
    assert item["dimensions"] == {"length_mm": 2700}
    assert item["dimensions_label"] == "2,7 м"


async def test_preview_missing_length_required_marks_invalid(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """Ветка 3: нет ни длины, ни типового размера, is_required → invalid."""
    product = await _make_product(session, "LEN-REQ")
    await _link_length_dimension(session, product, is_required=True, default_value=None)
    await session.commit()

    excel_buf = _make_excel([("LEN-REQ", 10, "", None)])
    resp = await client.post(
        "/api/stock/import/remainders/preview",
        files={"file": ("test.xlsx", excel_buf, XLSX_MIME)},
        data={"sheet_index": "0"},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["summary"]["invalid"] == 1
    item = body["items"][0]
    assert item["status"] == "invalid"
    assert any("Не указана длина" in e for e in item["errors"])
    assert any("Длина" in e for e in item["errors"])


async def test_preview_product_without_dimensions_stays_dimensionless(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """Крепёж (без привязок измерений): длина не требуется, dimensions=None."""
    await _make_product(session, "LEN-BOLT")
    await session.commit()

    excel_buf = _make_excel([("LEN-BOLT", 500, "", None)])
    resp = await client.post(
        "/api/stock/import/remainders/preview",
        files={"file": ("test.xlsx", excel_buf, XLSX_MIME)},
        data={"sheet_index": "0"},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["summary"]["invalid"] == 0
    item = body["items"][0]
    assert item["status"] == "valid"
    assert item["dimensions"] is None
    assert item["dimensions_label"] == "—"


# ─── Tests: Apply ──────────────────────────────────────────────────────────────


async def test_import_writes_dimensions_and_splits_balance_by_length(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """Применение пишет dimensions в транзакции; баланс разбит по длинам."""
    product = await _make_product(session, "LEN-APPLY")
    await _link_length_dimension(session, product, is_required=True, default_value=2700)
    location = await _make_location(session, "LEN-APPLY-LOC")
    await session.commit()

    excel_buf = _make_excel([
        ("LEN-APPLY", 100, "2,7", None),
        ("LEN-APPLY", 40, "0,9", None),
        ("LEN-APPLY", 10, "", None),  # пусто → типовой размер 2,7 м
    ])
    resp = await client.post(
        "/api/stock/import/remainders",
        files={"file": ("test.xlsx", excel_buf, XLSX_MIME)},
        data={
            "location_id": str(location.id),
            "quality_state": "good",
            "sheet_index": "0",
            "skip_invalid": "true",
            "clear_existing": "false",
        },
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["success"] is True
    assert body["imported_count"] == 3

    txs = (
        await session.scalars(
            select(StockTransaction).where(
                StockTransaction.id.in_(body["transaction_ids"])
            )
        )
    ).all()
    assert sorted(tx.dimensions["length_mm"] for tx in txs) == [900, 2700, 2700]

    balances = (
        await session.scalars(
            select(StockBalance).where(
                StockBalance.product_id == product.id,
                StockBalance.location_id == location.id,
            )
        )
    ).all()
    by_length = {bal.dimensions["length_mm"]: float(bal.balance_qty) for bal in balances}
    assert by_length == {2700: 110.0, 900: 40.0}

    await assert_no_invariants_violations(session, context="remainder-import-length")


async def test_import_dimensionless_product_writes_null_dimensions(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """Продукт без обязательных измерений импортируется с dimensions=NULL."""
    product = await _make_product(session, "LEN-NODIM")
    location = await _make_location(session, "LEN-NODIM-LOC")
    await session.commit()

    excel_buf = _make_excel([("LEN-NODIM", 25, "", None)])
    resp = await client.post(
        "/api/stock/import/remainders",
        files={"file": ("test.xlsx", excel_buf, XLSX_MIME)},
        data={
            "location_id": str(location.id),
            "quality_state": "good",
            "sheet_index": "0",
        },
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["imported_count"] == 1

    tx = await session.get(StockTransaction, body["transaction_ids"][0])
    assert tx is not None
    assert tx.dimensions is None

    bal = (
        await session.scalars(
            select(StockBalance).where(
                StockBalance.product_id == product.id,
                StockBalance.location_id == location.id,
            )
        )
    ).one()
    assert bal.dimensions is None
    assert float(bal.balance_qty) == 25.0


async def test_import_skips_row_without_length_and_default(
    client: AsyncClient,
    session: AsyncSession,
) -> None:
    """Строка без длины и типового размера (is_required) пропускается при skip_invalid."""
    good = await _make_product(session, "LEN-GOOD")
    bad = await _make_product(session, "LEN-BAD")
    await _link_length_dimension(session, bad, is_required=True, default_value=None)
    location = await _make_location(session, "LEN-SKIP-LOC")
    await session.commit()

    excel_buf = _make_excel([
        ("LEN-GOOD", 10, "1,8", None),
        ("LEN-BAD", 5, "", None),
    ])
    resp = await client.post(
        "/api/stock/import/remainders",
        files={"file": ("test.xlsx", excel_buf, XLSX_MIME)},
        data={
            "location_id": str(location.id),
            "sheet_index": "0",
            "skip_invalid": "true",
        },
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["success"] is True
    assert body["imported_count"] == 1
    assert any("Не указана длина" in e for e in body["errors"])

    tx = await session.get(StockTransaction, body["transaction_ids"][0])
    assert tx is not None
    assert tx.product_id == good.id
    assert tx.dimensions == {"length_mm": 1800}


# ─── Tests: Template ───────────────────────────────────────────────────────────


async def test_remainders_template_contains_length_column(
    session: AsyncSession,
) -> None:
    """Шаблон импорта остатков содержит колонку «Длина» и пример в метрах."""
    location = await _make_location(session, "LEN-TPL-LOC")
    await session.commit()

    template_bytes = await generate_remainders_template_for_location(
        session, location.id
    )
    wb = load_workbook(BytesIO(template_bytes))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert "Длина" in headers

    length_col = headers.index("Длина")
    example_values = [row[length_col].value for row in ws.iter_rows(min_row=2)]
    assert "2,7" in example_values

    # Шаблон парсится обратно нашим же парсером: длина распознаётся.
    _sheet, _total, items, _summary = await parse_remainders_excel(template_bytes)
    parsed = {it.sku: it.dimensions for it in items}
    assert parsed["ALS-1289"] == {"length_mm": 2700}
    assert parsed["361"] is None
